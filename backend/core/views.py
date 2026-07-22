import json
import logging
import mimetypes
import re
from collections import defaultdict
from io import BytesIO
from datetime import datetime, time, timedelta
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.db import transaction, IntegrityError
from django.db.models import Count, Q
from django.contrib.auth.hashers import make_password
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import AppUser, Lead, LeadStatusHistory, OnlineLead, LeadVisitDecision, LeadVisitDecisionHistory, ExcelImport, ExcelImportRow, DataAuditLog
from .auth import issue_tokens, verify_password, require_auth, decode_refresh
from .serializers import user_to_dict, serialize_leads, lead_to_dict, online_lead_to_dict, visit_decision_to_dict
from .utils import (
    STATUS_LABELS, STATUS_ORDER, NOTE_REQUIRED_STATUSES,
    clean_string, is_valid_status, empty_status_counts, normalize_excel_column,
    normalize_phone, parse_datetime_value, parse_date_range, day_range,
    apply_lead_search, excel_safe, format_dt, local_date_key,
)
from .telegram import send_telegram_message, line as tg_line, lead_info as tg_lead_info, user_name as tg_user_name, now_text as tg_now_text

logger = logging.getLogger(__name__)


def notify_telegram_after_commit(text):
    if not text:
        return
    transaction.on_commit(lambda: send_telegram_message(text), robust=True)



BRANCH_CHOICES = [
    'Niyozbosh',
    'Kids 1',
    'Kids 2',
    'Gulbahor',
    'Kids 3',
    'Kasblar',
    'Xalqobod',
    'Chinoz',
    'Olmozor',
    'Paxtazor',
    'Mevazor',
    'Dostobod',
    "Qorg'onchi",
    "Oqqo'rg'on",
    "Qo'shyog'och",
]

BRANCH_MANAGER_DISPLAY_NAMES = {
    # Tizimdagi eski "Olmozor" qiymatini buzmasdan foydalanuvchiga to‘g‘ri nom ko‘rsatamiz.
    'Olmozor': 'Olmazor',
}

BRANCH_ALIASES = {
    'niyozbosh': 'Niyozbosh', 'niyazbosh': 'Niyozbosh', 'niyoz bosh': 'Niyozbosh', 'niyaz bosh': 'Niyozbosh',
    'kids1': 'Kids 1', 'kids 1': 'Kids 1',
    'kids2': 'Kids 2', 'kids 2': 'Kids 2',
    'kids3': 'Kids 3', 'kids 3': 'Kids 3',
    'gulbahor': 'Gulbahor',
    'kasblar': 'Kasblar',
    'xalqobod': 'Xalqobod', 'xalqabod': 'Xalqobod',
    'chinoz': 'Chinoz',
    'olmazor': 'Olmozor', 'olmozor': 'Olmozor',
    'paxtazor': 'Paxtazor',
    'mevazor': 'Mevazor',
    'dostobod': 'Dostobod', 'do‘stobod': 'Dostobod', "do'stobod": 'Dostobod', 'do’stobod': 'Dostobod',
    'qorgonchi': "Qorg'onchi", "qorg'onchi": "Qorg'onchi", 'qorg’onchi': "Qorg'onchi",
    'qorgoncha': "Qorg'onchi", "qo'rg'oncha": "Qorg'onchi", 'qo’rg’oncha': "Qorg'onchi", 'qo‘rg‘oncha': "Qorg'onchi",
    'oqqorgon': "Oqqo'rg'on", "oqqo'rg'on": "Oqqo'rg'on", 'oqqo’rg’on': "Oqqo'rg'on", 'oqqo‘rg‘on': "Oqqo'rg'on",
    'qoshyogoch': "Qo'shyog'och", "qo'shyog'och": "Qo'shyog'och", 'qo’shyog’och': "Qo'shyog'och", 'qo‘shyog‘och': "Qo'shyog'och",
}

def normalize_branch_label(value):
    value = clean_string(value)
    if not value:
        return ''
    if '-' in value and value.split('-', 1)[0].strip().isdigit():
        value = value.split('-', 1)[1].strip()
    key = value.lower().replace('ʼ', "'").replace('’', "'").replace('‘', "'").strip()
    key = ' '.join(key.split())
    if key in BRANCH_ALIASES:
        return BRANCH_ALIASES[key]
    for branch in BRANCH_CHOICES:
        if branch.lower().replace('’', "'") == key:
            return branch
    return value


def normalize_branch_names(value):
    items = []
    if isinstance(value, (list, tuple)):
        raw_items = value
    else:
        raw_items = str(value or '').replace(';', ',').split(',')
    for raw in raw_items:
        branch = normalize_branch_label(raw)
        if branch and branch not in items:
            items.append(branch)
    return items


def branch_names_string(body):
    raw = body.get('branch_names')
    if raw is None:
        raw = body.get('branch_name') or body.get('branch') or body.get('filial')
    return ', '.join(normalize_branch_names(raw))


def user_branch_set(user):
    return set(normalize_branch_names(getattr(user, 'branch_name', '') or ''))


def manager_names_for_branches(branch_name_string):
    """Filiallar asosida ko‘rinadigan menenjer nomini yaratadi.

    Misol: "Olmozor" -> "Olmazor Menenjeri".
    Bir nechta filial bo‘lsa har biri alohida yoziladi.
    """
    names = []
    for branch in normalize_branch_names(branch_name_string):
        display_name = BRANCH_MANAGER_DISPLAY_NAMES.get(branch, branch)
        manager_name = f'{display_name} Menenjeri'
        if manager_name not in names:
            names.append(manager_name)
    return ', '.join(names)


def users_branch_overlap(user_a, user_b):
    a = user_branch_set(user_a)
    b = user_branch_set(user_b)
    return bool(a and b and a.intersection(b))


def known_branch_set(value):
    """Filial qiymatini faqat tizimdagi aniq filial nomlariga normallashtiradi."""
    return {branch for branch in normalize_branch_names(value) if branch in BRANCH_CHOICES}


def sale_branches_from_note(note):
    """Eski sotuv tarixidagi ``Filial: ...`` yozuvidan filialni topadi."""
    text = clean_string(note)
    if not text:
        return set()
    branches = set()
    # Izoh odatda: "izoh | Filial: Niyozbosh" ko‘rinishida saqlangan.
    for match in re.finditer(r'(?:^|[|;\n])\s*filial\s*:\s*([^|;\n]+)', text, flags=re.IGNORECASE):
        branches.update(known_branch_set(match.group(1)))
    return branches


def sale_history_map_for_leads(lead_ids):
    """Leadlar uchun sotuv vaqtida tanlangan filiallarni bitta so‘rovda oladi."""
    ids = [int(item) for item in lead_ids if item]
    if not ids:
        return {}
    result = defaultdict(set)
    rows = LeadStatusHistory.objects.filter(
        lead_id__in=ids,
        new_status='sale',
    ).only('lead_id', 'note', 'changed_at', 'id').order_by('-changed_at', '-id')
    for row in rows:
        if result.get(row.lead_id):
            continue
        branches = sale_branches_from_note(row.note)
        if branches:
            result[row.lead_id] = branches
    return result


def lead_sale_branch_set(lead, sale_history_branches=None):
    """Sotilgan lead qaysi filialga tegishli ekanini ishonchli aniqlaydi.

    Ustuvorlik:
    1) sotuv statusi tarixidagi ``Filial: ...``;
    2) leadning hozirgi ``branch_name`` qiymati;
    3) juda eski leadlarda operatorga biriktirilgan filiallar.

    Shu tartib eski Excel/import filiali ``branch_name``da qolib ketgan holatda ham
    keyin yaratilgan menenjerga sotuv leadini ko‘rsatadi.
    """
    history_branches = set(sale_history_branches or ())
    if sale_history_branches is None and getattr(lead, 'id', None):
        rows = LeadStatusHistory.objects.filter(
            lead_id=lead.id,
            new_status='sale',
        ).only('note').order_by('-changed_at', '-id')
        for row in rows:
            branches = sale_branches_from_note(row.note)
            if branches:
                history_branches = branches
                break
    if history_branches:
        return history_branches

    direct_branches = known_branch_set(getattr(lead, 'branch_name', '') or '')
    if direct_branches:
        return direct_branches

    operator = getattr(lead, 'assigned_operator', None)
    return user_branch_set(operator) if operator else set()


def filial_can_see_lead(filial_user, lead, sale_history_branches=None):
    # Huquq menenjer akkaunti yaratilgan vaqt yoki menenjer ID bo‘yicha emas,
    # sotuv vaqtida tanlangan filial bo‘yicha tekshiriladi.
    filial_branches = user_branch_set(filial_user)
    lead_branches = lead_sale_branch_set(lead, sale_history_branches)
    return bool(filial_branches and lead_branches and filial_branches.intersection(lead_branches))


def manager_visible_visit_decisions(user):
    """Menenjer filialiga tegishli eski va yangi nazorat yozuvlarini qaytaradi.

    Akkaunt o‘chirilishi soft-delete bo‘lgani uchun eski qarorlar bazada qoladi.
    Yangi menenjer ayni filialga biriktirilganda qarorlar decided_by ID bo‘yicha
    emas, lead filiali bo‘yicha topiladi. Bir lead uchun bir nechta eski yozuv
    bo‘lsa, Keldi yakuniy holati ustun; aks holda eng so‘nggi yozuv olinadi.
    """
    qs = LeadVisitDecision.objects.select_related(
        'lead', 'lead__assigned_operator', 'lead__boss',
        'decided_by', 'payment_done_by', 'payment_not_done_by', 'left_without_payment_by',
    ).order_by('-updated_at', '-id')
    items = list(qs)
    history_map = sale_history_map_for_leads([item.lead_id for item in items])
    selected = {}
    for item in items:
        lead = getattr(item, 'lead', None)
        if not lead or not filial_can_see_lead(user, lead, history_map.get(item.lead_id, set())):
            continue
        current = selected.get(item.lead_id)
        if current is None or (item.decision == 'arrived' and current.decision != 'arrived'):
            selected[item.lead_id] = item
    return sorted(selected.values(), key=lambda item: (item.updated_at, item.id), reverse=True)


def manager_existing_visit_decision(user, lead, *, for_update=False):
    """Filial bo‘yicha boshqariladigan bitta asosiy qarorni topadi."""
    qs = LeadVisitDecision.objects.select_related(
        'lead', 'decided_by', 'payment_done_by', 'payment_not_done_by', 'left_without_payment_by'
    ).filter(lead=lead)
    if for_update:
        qs = qs.select_for_update()
    items = list(qs.order_by('-updated_at', '-id'))
    if not items:
        return None
    # Keldi yakuniy qaror bo‘lgani uchun u har doim ustun turadi.
    for item in items:
        if item.decision == 'arrived':
            return item
    # Aks holda hozirgi menenjerning yozuvi yoki eng so‘nggi eski yozuv olinadi.
    for item in items:
        if item.decided_by_id == user.id:
            return item
    return items[0]

def visit_payment_status(item):
    """Menenjer aniq belgilagan to‘lov holatini qaytaradi."""
    if bool(getattr(item, 'payment_done', False)):
        return 'paid'
    if bool(getattr(item, 'payment_not_done', False)):
        return 'unpaid'
    if bool(getattr(item, 'left_without_payment', False)):
        return 'left_without_payment'
    return 'pending'


def visit_payment_label(item):
    return {
        'paid': 'To‘lov qildi',
        'unpaid': 'To‘lov qilmadi',
        'left_without_payment': 'Keldi, to‘lov qilmasdan ketdi',
        'pending': 'Belgilanmagan',
    }[visit_payment_status(item)]


def visit_payment_status_at(item):
    status = visit_payment_status(item)
    if status == 'paid':
        return getattr(item, 'payment_done_at', None)
    if status == 'unpaid':
        return getattr(item, 'payment_not_done_at', None)
    if status == 'left_without_payment':
        return getattr(item, 'left_without_payment_at', None)
    return None


def visit_payment_status_by_name(item):
    status = visit_payment_status(item)
    user = None
    if status == 'paid':
        user = getattr(item, 'payment_done_by', None)
    elif status == 'unpaid':
        user = getattr(item, 'payment_not_done_by', None)
    elif status == 'left_without_payment':
        user = getattr(item, 'left_without_payment_by', None)
    return (user.full_name or user.username) if user else ''

def status_text(value):
    return STATUS_LABELS.get(value, value or '')


def calc_conversion(sale_count, total_count):
    """Konversiya foizini hisoblaydi va hech qachon 100% dan oshmasligini ta'minlaydi."""
    if not total_count:
        return 0.0
    value = (sale_count / total_count) * 100
    return round(min(value, 100), 1)


def json_body(request):
    if not request.body:
        return {}
    try:
        return json.loads(request.body.decode('utf-8'))
    except Exception:
        return {}


def ok(data, status=200):
    return JsonResponse(data, status=status, safe=not isinstance(data, list))


def duplicate_name_key(value):
    return ' '.join(clean_string(value).lower().split())


def duplicate_phone_key(value):
    normalized = normalize_phone(value)
    digits = ''.join(ch for ch in normalized if ch.isdigit())
    if not digits:
        return ''
    # +99890xxxxxxx, 909xxxxxx, 8(90)xxxxxxx ko‘rinishlarini bir xil solishtirish uchun
    # oxirgi 9 ta raqam yetarli bo‘ladi.
    return digits[-9:] if len(digits) >= 9 else digits


def lead_phone_keys(*phones):
    keys = []
    for phone in phones:
        key = duplicate_phone_key(phone)
        if key and key not in keys:
            keys.append(key)
    return keys


def duplicate_lookup_keys(full_name, phone1='', phone2='', phone3=''):
    """
    Dublikat kaliti: normallashtirilgan F.I.SH + har bir telefonning oxirgi 9 raqami.
    Excel ichida ham, umumiy bazada ham bir xil odam qayta qo‘shilmasligi uchun ishlatiladi.
    """
    name_key = duplicate_name_key(full_name)
    phone_keys = lead_phone_keys(phone1, phone2, phone3)
    if not name_key or not phone_keys:
        return []
    return [(name_key, phone_key) for phone_key in phone_keys]


def add_lead_to_duplicate_index(index, lead):
    for key in duplicate_lookup_keys(lead.full_name, lead.phone1, lead.phone2, lead.phone3):
        bucket = index[key]
        if not any(getattr(item, 'id', None) == getattr(lead, 'id', None) for item in bucket):
            bucket.append(lead)


def build_duplicate_lead_index():
    """
    Umumiy Lead bazadagi barcha leadlarni Python tarafda normallashtirib index qiladi.
    Bu full_name__iexact yoki telefon formatidagi bo‘sh joy/.0 farqlari sabab dublikat o‘tib ketishini to‘xtatadi.
    """
    index = defaultdict(list)
    qs = Lead.objects.select_related('assigned_operator', 'boss').only(
        'id', 'full_name', 'phone1', 'phone2', 'phone3', 'current_status', 'created_at',
        'assigned_operator__id', 'assigned_operator__full_name', 'assigned_operator__username',
        'boss__id', 'boss__full_name', 'boss__username',
    ).order_by('-created_at', '-id')
    for lead in qs.iterator(chunk_size=1000):
        add_lead_to_duplicate_index(index, lead)
    return index


def find_duplicate_leads_in_index(index, full_name, phone1='', phone2='', phone3='', limit=20):
    matches = []
    seen_ids = set()
    for key in duplicate_lookup_keys(full_name, phone1, phone2, phone3):
        for lead in index.get(key, []):
            lead_id = getattr(lead, 'id', None)
            if lead_id in seen_ids:
                continue
            seen_ids.add(lead_id)
            matches.append(lead)
            if len(matches) >= limit:
                return matches
    return matches


def find_existing_general_leads(full_name, phone1='', phone2='', phone3='', limit=20):
    """
    Umumiy Lead bazadan F.I.SH + kamida bitta telefon raqami bir xil leadlarni topadi.
    Bossga bog‘lamaymiz: umumiy bazada bor bo‘lsa, qayta qo‘shilmaydi.
    """
    duplicate_index = build_duplicate_lead_index()
    return find_duplicate_leads_in_index(duplicate_index, full_name, phone1, phone2, phone3, limit=limit)


def duplicate_leads_payload(leads):
    items = []
    for lead in leads:
        operator = lead.assigned_operator if getattr(lead, 'assigned_operator_id', None) else None
        boss = lead.boss if getattr(lead, 'boss_id', None) else None
        items.append({
            'id': lead.id,
            'full_name': lead.full_name or '',
            'phone1': lead.phone1 or '',
            'phone2': lead.phone2 or '',
            'phone3': lead.phone3 or '',
            'operator_name': ((operator.full_name or operator.username) if operator else ''),
            'boss_name': ((boss.full_name or boss.username) if boss else ''),
            'current_status': lead.current_status or '',
            'status_label': status_text(lead.current_status),
            'created_at': format_dt(lead.created_at),
        })
    return items


def duplicate_block_detail(duplicates):
    payload = duplicate_leads_payload(duplicates)
    lines = [f"Bu lead umumiy bazada oldin biriktirilgan. Jami: {len(payload)} ta."]
    for idx, item in enumerate(payload[:5], 1):
        phones = ', '.join([p for p in [item.get('phone1'), item.get('phone2'), item.get('phone3')] if p]) or '-'
        operator_name = item.get('operator_name') or '-'
        lines.append(f"{idx}) {item.get('full_name') or '-'} | Tel: {phones} | Operator: {operator_name}")
    if len(payload) > 5:
        lines.append(f"Yana {len(payload) - 5} ta dublikat bor.")
    return '\n'.join(lines)


def duplicate_block_response(duplicates):
    return ok({
        'detail': duplicate_block_detail(duplicates),
        'duplicate_count': len(duplicates),
        'duplicate_leads': duplicate_leads_payload(duplicates),
    }, status=409)


MONTH_NAMES_UZ = {
    1: 'Yanvar', 2: 'Fevral', 3: 'Mart', 4: 'Aprel', 5: 'May', 6: 'Iyun',
    7: 'Iyul', 8: 'Avgust', 9: 'Sentabr', 10: 'Oktabr', 11: 'Noyabr', 12: 'Dekabr',
}


def add_months(date_value, delta):
    month_index = (date_value.year * 12 + (date_value.month - 1)) + delta
    year = month_index // 12
    month = month_index % 12 + 1
    return datetime(year, month, 1).date()


def parse_month_key(value):
    text = clean_string(value)
    if len(text) != 7 or '-' not in text:
        return None
    try:
        year, month = [int(part) for part in text.split('-', 1)]
        if month < 1 or month > 12:
            return None
        return datetime(year, month, 1).date()
    except Exception:
        return None


def month_key_from_date(date_value):
    return f'{date_value.year:04d}-{date_value.month:02d}'


def local_month_key(value=None):
    value = value or timezone.now()
    if isinstance(value, str):
        value = parse_datetime_value(value) or timezone.now()
    local_value = timezone.localtime(value) if timezone.is_aware(value) else value
    return f'{local_value.year:04d}-{local_value.month:02d}'


def month_label(month_key):
    parsed = parse_month_key(month_key)
    if not parsed:
        return month_key or ''
    return f"{MONTH_NAMES_UZ.get(parsed.month, parsed.month)} {parsed.year}"


def month_period_from_params(params, default_months=12):
    params = params or {}
    end_month = parse_month_key(params.get('end_month')) or parse_month_key(params.get('month')) or timezone.localdate().replace(day=1)
    start_month = parse_month_key(params.get('start_month'))
    if not start_month:
        try:
            count = int(params.get('months') or default_months)
        except Exception:
            count = default_months
        count = max(1, min(count, 36))
        start_month = add_months(end_month, -(count - 1))
    if start_month > end_month:
        start_month, end_month = end_month, start_month
    month_keys = []
    cursor = start_month
    while cursor <= end_month:
        month_keys.append(month_key_from_date(cursor))
        cursor = add_months(cursor, 1)
    end_next = add_months(end_month, 1)
    start_dt = timezone.make_aware(datetime.combine(start_month, time.min))
    end_dt = timezone.make_aware(datetime.combine(end_next - timedelta(days=1), time.max))
    return {
        'start_month': month_key_from_date(start_month),
        'end_month': month_key_from_date(end_month),
        'keys': month_keys,
        'start_dt': start_dt,
        'end_dt': end_dt,
    }


def empty_month_row(key):
    return {
        'month': key,
        'month_label': month_label(key),
        'assigned_leads': 0,
        'sale': 0,
        'otkaz': 0,
        'wrong_number': 0,
        'open_number': 0,
        'advice': 0,
        'other': 0,
        'not_answered': 0,
        'actions_total': 0,
        'online_submitted': 0,
        'online_assigned': 0,
        'arrived': 0,
        'not_arrived': 0,
        'payment_done': 0,
        'payment_not_done': 0,
        'left_without_payment': 0,
        'payment_pending': 0,
    }


def build_monthly_archive_for_operators(op_ids, params=None, *, boss_user=None, include_online_submitted=False, include_decisions=False):
    """
    Oyma-oy arxiv LeadStatusHistory.changed_at bo‘yicha hisoblanadi.
    Shu sabab eski oy natijasi yangi oyga aralashmaydi; oy almashganda yangi oy avtomatik ochiladi.
    """
    period = month_period_from_params(params or {})
    rows = {key: empty_month_row(key) for key in period['keys']}
    if not op_ids:
        return [rows[key] for key in reversed(period['keys'])]

    assigned_qs = Lead.objects.filter(
        assigned_operator_id__in=op_ids,
        created_at__range=(period['start_dt'], period['end_dt']),
    ).only('created_at')
    for lead in assigned_qs.iterator(chunk_size=1000):
        key = local_month_key(lead.created_at)
        if key in rows:
            rows[key]['assigned_leads'] += 1

    hist_qs = LeadStatusHistory.objects.filter(
        changed_by_id__in=op_ids,
        changed_at__range=(period['start_dt'], period['end_dt']),
    ).only('changed_at', 'new_status')
    for item in hist_qs.iterator(chunk_size=1000):
        key = local_month_key(item.changed_at)
        if key not in rows:
            continue
        rows[key]['actions_total'] += 1
        if item.new_status in rows[key]:
            rows[key][item.new_status] += 1

    online_assigned_qs = OnlineLead.objects.filter(
        assigned_operator_id__in=op_ids,
        assigned_at__range=(period['start_dt'], period['end_dt']),
    ).only('assigned_at')
    for item in online_assigned_qs.iterator(chunk_size=1000):
        key = local_month_key(item.assigned_at)
        if key in rows:
            rows[key]['online_assigned'] += 1

    if include_online_submitted:
        online_submitted_qs = OnlineLead.objects.filter(
            submitted_at__range=(period['start_dt'], period['end_dt']),
        ).only('submitted_at')
        if boss_user is not None:
            online_submitted_qs = online_submitted_qs.filter(assigned_boss=boss_user)
        for item in online_submitted_qs.iterator(chunk_size=1000):
            key = local_month_key(item.submitted_at)
            if key in rows:
                rows[key]['online_submitted'] += 1

    if include_decisions:
        decisions_qs = LeadVisitDecision.objects.filter(
            lead__assigned_operator_id__in=op_ids,
            updated_at__range=(period['start_dt'], period['end_dt']),
        ).only('updated_at', 'decision', 'payment_done', 'payment_not_done', 'left_without_payment')
        for item in decisions_qs.iterator(chunk_size=1000):
            key = local_month_key(item.updated_at)
            if key not in rows:
                continue
            rows[key]['arrived' if item.decision == 'arrived' else 'not_arrived'] += 1
            payment_key = {
                'paid': 'payment_done',
                'unpaid': 'payment_not_done',
                'left_without_payment': 'left_without_payment',
                'pending': 'payment_pending',
            }[visit_payment_status(item)]
            rows[key][payment_key] += 1

    for row in rows.values():
        total_base = row['assigned_leads'] or row['actions_total']
        row['conversion'] = calc_conversion(row['sale'], total_base)
    return [rows[key] for key in reversed(period['keys'])]


def health(request):
    return ok({'status': 'ok', 'message': 'Django backend ishlayapti'})


@require_auth('admin')
def telegram_test(request):
    result = send_telegram_message(
        '✅ <b>Telegram notification test</b>\n'
        + tg_line('Vaqt', tg_now_text())
        + tg_line('Admin', tg_user_name(request.app_user))
        + 'Aziz Academy Operators CRM dan test xabar.'
    )
    return ok({'sent': bool(result.get('ok')), 'result': result})


@csrf_exempt
@require_http_methods(['POST'])
def login_view(request):
    body = json_body(request)
    username = clean_string(body.get('username'))
    password = str(body.get('password') or '')
    user = AppUser.objects.filter(username=username, is_active=True).first()
    if not user or not verify_password(password, user.password_hash):
        return ok({'non_field_errors': ['Login yoki parol xato.'], 'detail': 'Login yoki parol xato.'}, status=400)
    tokens = issue_tokens(user)
    return ok({**tokens, 'user': user_to_dict(user)})


@csrf_exempt
@require_http_methods(['POST'])
def refresh_view(request):
    body = json_body(request)
    try:
        payload = decode_refresh(body.get('refresh'))
        user = AppUser.objects.get(id=payload.get('userId'), is_active=True)
        return ok({'access': issue_tokens(user)['access']})
    except Exception:
        return ok({'detail': 'Refresh token noto‘g‘ri yoki muddati tugagan.'}, status=401)


@require_auth()
def me_view(request):
    return ok(user_to_dict(request.app_user))


def list_users(role, boss_id=None):
    qs = AppUser.objects.select_related('boss').filter(role=role, is_active=True)
    if boss_id is not None:
        qs = qs.filter(boss_id=boss_id)
    return [user_to_dict(u) for u in qs.order_by('-id')]


def create_user(request, role, boss_id=None):
    body = json_body(request)
    username = clean_string(body.get('username'))
    password = str(body.get('password') or '')
    full_name = clean_string(body.get('full_name'))
    phone = '' if role in ('operator', 'filial_rahbari') else clean_string(body.get('phone'))
    branch_name = branch_names_string(body)
    if role in ('operator', 'filial_rahbari'):
        if role == 'filial_rahbari':
            full_name = full_name or manager_names_for_branches(branch_name) or username
        else:
            full_name = full_name or username
        if not username or not password:
            return ok({'detail': 'Login va parol kiritish shart.'}, status=400)
        if not branch_name:
            return ok({'branch_name': ['Kamida bitta filial tanlash shart.'], 'detail': 'Kamida bitta filial tanlash shart.'}, status=400)
    elif not username or not password or not full_name:
        return ok({'detail': 'Ism, login va parol kiritish shart.'}, status=400)
    # O‘chirilgan akkaunt bazadan butunlay o‘chirilmaydi: tarix, lead va
    # statistika yo‘qolmasligi uchun u faqat is_active=False qilinadi. Shu
    # sabab avval ishlatilgan login bilan qayta yaratishda eski yozuvni qayta
    # faollashtiramiz. Bu login unique xatosini bartaraf qiladi va eski
    # menenjer qayta ochilsa uning tarixiy qarorlari ham saqlanib qoladi.
    inactive_user = AppUser.objects.filter(username=username, is_active=False).first()
    if inactive_user:
        if inactive_user.role != role:
            return ok({
                'username': ['Bu login boshqa turdagi o‘chirilgan akkauntga tegishli. Boshqa login kiriting.'],
                'detail': 'Bu login boshqa turdagi o‘chirilgan akkauntga tegishli.',
            }, status=400)
        old = user_to_dict(inactive_user)
        inactive_user.password_hash = make_password(password)
        inactive_user.full_name = full_name
        inactive_user.phone = phone
        inactive_user.role = role
        inactive_user.boss_id = boss_id
        inactive_user.branch_name = branch_name
        inactive_user.is_active = bool(body.get('is_active', True))
        inactive_user.deactivated_at = None if inactive_user.is_active else (inactive_user.deactivated_at or timezone.now())
        inactive_user.updated_at = timezone.now()
        inactive_user.save(update_fields=[
            'password_hash', 'full_name', 'phone', 'role', 'boss_id',
            'branch_name', 'is_active', 'deactivated_at', 'updated_at',
        ])
        DataAuditLog.objects.create(
            actor=request.app_user,
            entity_type='app_user',
            entity_id=inactive_user.id,
            action='reactivated',
            old_data=old,
            new_data=user_to_dict(inactive_user),
        )
        payload = user_to_dict(inactive_user)
        payload['detail'] = 'O‘chirilgan akkaunt qayta faollashtirildi.'
        payload['reactivated'] = True
        return ok(payload, status=201)

    if AppUser.objects.filter(username=username).exists():
        return ok({'username': ['Bu login allaqachon mavjud.']}, status=400)
    try:
        user = AppUser.objects.create(
            username=username,
            password_hash=make_password(password),
            full_name=full_name,
            phone=phone,
            role=role,
            boss_id=boss_id,
            branch_name=branch_name,
            is_active=bool(body.get('is_active', True)),
        )
    except IntegrityError as exc:
        message = str(exc).lower()
        if 'username' in message:
            return ok({'username': ['Bu login allaqachon mavjud.'], 'detail': 'Bu login allaqachon mavjud.'}, status=400)
        if 'branch' in message or 'filial' in message:
            return ok({
                'branch_name': ['Filial bo‘yicha eski unique cheklov qolgan. Backend qayta deploy qilinsa init_db uni avtomatik olib tashlaydi.'],
                'detail': 'Bu filialga boshqa operator biriktirilgan bo‘lsa ham yaratish mumkin. Backendni qayta deploy qiling va qayta urinib ko‘ring.',
            }, status=400)
        return ok({'detail': 'Foydalanuvchi yaratishda bazada xatolik yuz berdi.'}, status=400)
    return ok(user_to_dict(user), status=201)


def update_user(request, user_id, role, boss_id=None):
    qs = AppUser.objects.filter(id=user_id, role=role)
    if boss_id is not None:
        qs = qs.filter(boss_id=boss_id)
    user = qs.first()
    if not user:
        return ok({'detail': 'Foydalanuvchi topilmadi.'}, status=404)
    old = user_to_dict(user)
    body = json_body(request)
    username = clean_string(body.get('username') or user.username)
    if username != user.username and AppUser.objects.filter(username=username).exclude(id=user.id).exists():
        return ok({'username': ['Bu login allaqachon mavjud.']}, status=400)
    user.username = username
    if 'full_name' in body:
        user.full_name = clean_string(body.get('full_name'))
    if 'phone' in body:
        user.phone = '' if role in ('operator', 'filial_rahbari') else clean_string(body.get('phone'))
    if 'branch_names' in body or 'branch_name' in body or 'branch' in body or 'filial' in body:
        normalized_branches = branch_names_string(body)
        if role in ('operator', 'filial_rahbari') and not normalized_branches:
            return ok({'branch_name': ['Kamida bitta filial tanlash shart.'], 'detail': 'Kamida bitta filial tanlash shart.'}, status=400)
        user.branch_name = normalized_branches
    if 'is_active' in body:
        user.is_active = bool(body.get('is_active'))
        if user.is_active:
            user.deactivated_at = None
        elif not user.deactivated_at:
            user.deactivated_at = timezone.now()
    if body.get('password'):
        user.password_hash = make_password(str(body.get('password')))
    if role == 'filial_rahbari' and not user.full_name:
        user.full_name = manager_names_for_branches(user.branch_name) or user.username
    user.updated_at = timezone.now()
    user.save()
    DataAuditLog.objects.create(actor=request.app_user, entity_type='app_user', entity_id=user.id, action='updated', old_data=old, new_data=user_to_dict(user))
    return ok(user_to_dict(user))


def transfer_operator_data(from_user, to_user, actor):
    """from_user operatoriga tegishli barcha ma'lumotlarni (leadlar, online leadlar,
    excel importlar, status tarixi va h.k.) to_user operatoriga o'tkazadi."""
    with transaction.atomic():
        leads_count = Lead.objects.filter(assigned_operator=from_user).update(assigned_operator=to_user, updated_at=timezone.now())
        online_leads_count = OnlineLead.objects.filter(assigned_operator=from_user).update(assigned_operator=to_user)
        history_count = LeadStatusHistory.objects.filter(changed_by=from_user).update(changed_by=to_user)
        excel_imports_count = ExcelImport.objects.filter(assigned_operator=from_user).update(assigned_operator=to_user)
        DataAuditLog.objects.create(
            actor=actor,
            entity_type='app_user',
            entity_id=from_user.id,
            action='data_transferred',
            old_data={'from_operator_id': from_user.id, 'from_operator_name': from_user.full_name or from_user.username},
            new_data={
                'to_operator_id': to_user.id,
                'to_operator_name': to_user.full_name or to_user.username,
                'leads_transferred': leads_count,
                'online_leads_transferred': online_leads_count,
                'history_transferred': history_count,
                'excel_imports_transferred': excel_imports_count,
            },
        )
    return {
        'leads_transferred': leads_count,
        'online_leads_transferred': online_leads_count,
        'history_transferred': history_count,
        'excel_imports_transferred': excel_imports_count,
    }


def delete_user(request, user_id, role, boss_id=None):
    qs = AppUser.objects.filter(id=user_id, role=role)
    if boss_id is not None:
        qs = qs.filter(boss_id=boss_id)
    user = qs.first()
    if not user:
        return ok({'detail': 'Foydalanuvchi topilmadi.'}, status=404)

    transfer_summary = None
    if role == 'operator':
        raw_transfer_id = request.GET.get('transfer_to')
        if not raw_transfer_id:
            body = json_body(request)
            raw_transfer_id = body.get('transfer_to')
        try:
            transfer_to_id = int(raw_transfer_id) if raw_transfer_id else None
        except (TypeError, ValueError):
            transfer_to_id = None
        if not transfer_to_id:
            return ok({'detail': 'Operatorni o‘chirish uchun ma’lumotlar o‘tkaziladigan operatorni tanlash shart.', 'code': 'transfer_required'}, status=400)
        if transfer_to_id == user.id:
            return ok({'detail': 'O‘chirilayotgan operatorning o‘ziga ma’lumot o‘tkazib bo‘lmaydi.'}, status=400)
        target_qs = AppUser.objects.filter(id=transfer_to_id, role='operator', is_active=True)
        if boss_id is not None:
            target_qs = target_qs.filter(boss_id=boss_id)
        target_user = target_qs.first()
        if not target_user:
            return ok({'detail': 'Tanlangan operator topilmadi.'}, status=404)
        transfer_summary = transfer_operator_data(user, target_user, request.app_user)

    old = user_to_dict(user)
    user.is_active = False
    user.deactivated_at = user.deactivated_at or timezone.now()
    user.updated_at = timezone.now()
    user.save(update_fields=['is_active', 'deactivated_at', 'updated_at'])
    DataAuditLog.objects.create(actor=request.app_user, entity_type='app_user', entity_id=user.id, action='deactivated', old_data=old, new_data=user_to_dict(user))
    if transfer_summary is not None:
        return ok({'detail': 'Operator o‘chirildi va ma’lumotlari o‘tkazildi.', **transfer_summary}, status=200)
    return HttpResponse(status=204)


@csrf_exempt
@require_auth('admin')
def admin_bosses(request, user_id=None):
    if request.method == 'GET':
        return ok(list_users('boss'))
    if request.method == 'POST':
        return create_user(request, 'boss')
    if request.method in ('PATCH', 'PUT'):
        return update_user(request, user_id, 'boss')
    if request.method == 'DELETE':
        return delete_user(request, user_id, 'boss')
    return ok({'detail': 'Method not allowed'}, status=405)


@csrf_exempt
@require_auth('admin')
def admin_filial_rahbarlari(request, user_id=None):
    if request.method == 'GET':
        return ok(list_users('filial_rahbari'))
    if request.method == 'POST':
        return create_user(request, 'filial_rahbari')
    if request.method in ('PATCH', 'PUT'):
        return update_user(request, user_id, 'filial_rahbari')
    if request.method == 'DELETE':
        return delete_user(request, user_id, 'filial_rahbari')
    return ok({'detail': 'Method not allowed'}, status=405)



@csrf_exempt
@require_auth('admin')
def admin_directors(request, user_id=None):
    if request.method == 'GET':
        qs = AppUser.objects.filter(role__in=('director', 'director_deputy'), is_active=True).order_by('-id')
        return ok([user_to_dict(u) for u in qs])
    if request.method == 'POST':
        body = json_body(request)
        role = body.get('role') if body.get('role') in ('director', 'director_deputy') else 'director'
        return create_user(request, role)
    user = AppUser.objects.filter(id=user_id, role__in=('director', 'director_deputy')).first()
    if not user:
        return ok({'detail': 'Foydalanuvchi topilmadi.'}, status=404)
    if request.method in ('PATCH', 'PUT'):
        return update_user(request, user_id, user.role)
    if request.method == 'DELETE':
        return delete_user(request, user_id, user.role)
    return ok({'detail': 'Method not allowed'}, status=405)



@csrf_exempt
@require_auth('admin')
def admin_users(request, user_id=None):
    allowed_roles = ['boss', 'operator', 'filial_rahbari']
    if request.method == 'GET':
        qs = AppUser.objects.select_related('boss').filter(role__in=allowed_roles, is_active=True).order_by('role', '-id')
        return ok([user_to_dict(u) for u in qs])

    user = AppUser.objects.filter(id=user_id, role__in=allowed_roles).first()
    if not user:
        return ok({'detail': 'Foydalanuvchi topilmadi.'}, status=404)

    if request.method in ('PATCH', 'PUT'):
        return update_user(request, user_id, user.role)

    if request.method == 'DELETE':
        return delete_user(request, user_id, user.role)

    return ok({'detail': 'Method not allowed'}, status=405)

@csrf_exempt
@require_auth('boss')
def boss_operators(request, user_id=None):
    if request.method == 'GET':
        return ok(list_users('operator', request.app_user.id))
    if request.method == 'POST':
        return create_user(request, 'operator', request.app_user.id)
    if request.method in ('PATCH', 'PUT'):
        return update_user(request, user_id, 'operator', request.app_user.id)
    if request.method == 'DELETE':
        return delete_user(request, user_id, 'operator', request.app_user.id)
    return ok({'detail': 'Method not allowed'}, status=405)


@csrf_exempt
@require_auth('admin')
def admin_staff(request, user_id=None):
    """Admin panelida barcha faol operator va menenjer akkauntlarini boshqarish."""
    allowed_roles = ('operator', 'filial_rahbari')
    if request.method == 'GET':
        qs = AppUser.objects.select_related('boss').filter(role__in=allowed_roles, is_active=True).order_by('role', 'full_name', 'username')
        return ok([user_to_dict(user) for user in qs])

    user = AppUser.objects.select_related('boss').filter(id=user_id, role__in=allowed_roles, is_active=True).first()
    if not user:
        return ok({'detail': 'Operator yoki menenjer topilmadi.'}, status=404)

    if request.method in ('PATCH', 'PUT'):
        body = json_body(request)
        old = user_to_dict(user)
        username = clean_string(body.get('username') or user.username)
        full_name = clean_string(body.get('full_name')) if 'full_name' in body else user.full_name
        if not username:
            return ok({'username': ['Login bo‘sh bo‘lishi mumkin emas.']}, status=400)
        if AppUser.objects.filter(username=username).exclude(id=user.id).exists():
            return ok({'username': ['Bu login allaqachon mavjud.']}, status=400)
        if not full_name:
            return ok({'full_name': ['Ism-familiya kiritish shart.']}, status=400)

        if 'branch_names' in body or 'branch_name' in body or 'branch' in body or 'filial' in body:
            branch_name = branch_names_string(body)
            if not branch_name:
                return ok({'branch_name': ['Kamida bitta filial tanlash shart.']}, status=400)
            user.branch_name = branch_name

        user.username = username
        user.full_name = full_name
        if body.get('password'):
            user.password_hash = make_password(str(body.get('password')))
        user.updated_at = timezone.now()
        user.save()
        DataAuditLog.objects.create(
            actor=request.app_user,
            entity_type='app_user',
            entity_id=user.id,
            action='admin_staff_updated',
            old_data=old,
            new_data=user_to_dict(user),
        )
        return ok(user_to_dict(user))

    if request.method == 'DELETE':
        old = user_to_dict(user)
        user.is_active = False
        user.deactivated_at = user.deactivated_at or timezone.now()
        user.updated_at = timezone.now()
        user.save(update_fields=['is_active', 'deactivated_at', 'updated_at'])
        DataAuditLog.objects.create(
            actor=request.app_user,
            entity_type='app_user',
            entity_id=user.id,
            action='admin_staff_deactivated',
            old_data=old,
            new_data=user_to_dict(user),
        )
        return ok({'detail': 'Akkaunt o‘chirildi.'})

    return ok({'detail': 'Method not allowed'}, status=405)


@csrf_exempt
@require_http_methods(['POST'])
@require_auth('admin')
def admin_staff_bulk_delete(request):
    body = json_body(request)
    raw_ids = body.get('ids') or []
    if not isinstance(raw_ids, list):
        return ok({'detail': 'ids ro‘yxat ko‘rinishida yuborilishi kerak.'}, status=400)
    ids = []
    for value in raw_ids:
        try:
            item_id = int(value)
        except (TypeError, ValueError):
            continue
        if item_id not in ids:
            ids.append(item_id)
    if not ids:
        return ok({'detail': 'O‘chirish uchun kamida bitta akkaunt tanlang.'}, status=400)

    users = list(AppUser.objects.filter(id__in=ids, role__in=('operator', 'filial_rahbari'), is_active=True))
    if not users:
        return ok({'detail': 'Tanlangan faol akkauntlar topilmadi.'}, status=404)

    now = timezone.now()
    with transaction.atomic():
        for user in users:
            old = user_to_dict(user)
            user.is_active = False
            user.deactivated_at = user.deactivated_at or now
            user.updated_at = now
            user.save(update_fields=['is_active', 'deactivated_at', 'updated_at'])
            DataAuditLog.objects.create(
                actor=request.app_user,
                entity_type='app_user',
                entity_id=user.id,
                action='admin_staff_bulk_deactivated',
                old_data=old,
                new_data=user_to_dict(user),
            )
    return ok({'detail': f'{len(users)} ta akkaunt o‘chirildi.', 'deleted_count': len(users)})


def base_lead_qs():
    return Lead.objects.select_related('assigned_operator', 'boss', 'uploaded_by').all()


def get_one_lead(lead_id):
    lead = base_lead_qs().filter(id=lead_id).first()
    if not lead:
        return None
    online = OnlineLead.objects.select_related('assigned_operator').filter(created_lead_id=lead_id).order_by('-submitted_at').first()
    history = list(LeadStatusHistory.objects.select_related('changed_by').filter(lead_id=lead_id).order_by('-changed_at', '-id'))
    decisions = list(LeadVisitDecision.objects.select_related('lead', 'decided_by', 'payment_done_by', 'payment_not_done_by', 'left_without_payment_by').filter(lead_id=lead_id).order_by('-updated_at'))
    return lead_to_dict(lead, history, online, decisions)


@require_auth('operator')
def operator_leads(request):
    qs = base_lead_qs().filter(assigned_operator=request.app_user)
    status = clean_string(request.GET.get('current_status'))
    if status:
        qs = qs.filter(current_status=status)
    qs = apply_lead_search(qs, request.GET.get('search'))
    operator_branch_names = normalize_branch_names(getattr(request.app_user, 'branch_name', '') or '')
    return ok({
        'results': serialize_leads(qs.order_by('-updated_at', '-id'), include_visit_decisions=False),
        'operator_branch_names': operator_branch_names,
        'operator_branch_name': ', '.join(operator_branch_names),
    })


@csrf_exempt
@require_auth('operator')
def operator_incoming_leads(request):
    """Operator o'zi 'Lead qo'shish' orqali qo'shgan (kiruvchi qo'ng'iroq) leadlar ro'yxati."""
    qs = base_lead_qs().filter(assigned_operator=request.app_user, is_manual_entry=True)
    status = clean_string(request.GET.get('current_status'))
    if status:
        qs = qs.filter(current_status=status)
    qs = apply_lead_search(qs, request.GET.get('search'))
    return ok(serialize_leads(qs.order_by('-updated_at', '-id'), include_visit_decisions=False))


@csrf_exempt
@require_auth('operator')
def operator_create_lead(request):
    """Operator 'Lead qo'shish' tugmasi orqali qo'lda yangi lead (kiruvchi qo'ng'iroq) yaratadi."""
    if request.method != 'POST':
        return ok({'detail': 'Method not allowed'}, status=405)
    body = json_body(request)
    full_name = clean_string(body.get('full_name'))
    phone1 = normalize_phone(body.get('phone1') or body.get('phone') or '')
    phone2 = normalize_phone(body.get('phone2') or '')
    subject = clean_string(body.get('subject'))
    if not full_name or not phone1:
        return ok({'detail': 'Ism Familya va Nomer 1 kiritish shart.'}, status=400)

    operator = request.app_user
    lead = Lead.objects.create(
        full_name=full_name,
        phone1=phone1,
        phone2=phone2,
        subject=subject,
        assigned_operator=operator,
        boss_id=operator.boss_id,
        uploaded_by=operator,
        current_status='new',
        is_manual_entry=True,
    )
    LeadStatusHistory.objects.create(lead=lead, old_status='', new_status='new', changed_by=operator, note='Operator tomonidan qo‘shildi (kiruvchi qo‘ng‘iroq)')
    try:
        DataAuditLog.objects.create(
            actor=operator,
            entity_type='lead',
            entity_id=lead.id,
            action='manual_created',
            old_data={},
            new_data={'full_name': full_name, 'phone1': phone1, 'phone2': phone2, 'subject': subject},
        )
    except Exception:
        pass
    msg = (
        '📞 <b>Kiruvchi qo‘ng‘iroq (qo‘lda qo‘shilgan lead)</b>\n'
        + tg_line('Vaqt', tg_now_text())
        + tg_line('Operator', tg_user_name(operator))
        + tg_line('F.I.O', full_name)
        + tg_line('Telefon 1', phone1)
        + (tg_line('Telefon 2', phone2) if phone2 else '')
        + (tg_line('Fan', subject) if subject else '')
    )
    notify_telegram_after_commit(msg)
    return ok(get_one_lead(lead.id), status=201)


@csrf_exempt
@require_auth('operator')
def operator_update_lead_name(request, lead_id):
    """Operator o'ziga biriktirilgan leadning F.I.O sini tahrirlaydi."""
    if request.method not in ('PATCH', 'PUT'):
        return ok({'detail': 'Method not allowed'}, status=405)
    lead = Lead.objects.filter(id=lead_id, assigned_operator=request.app_user).first()
    if not lead:
        return ok({'detail': 'Bu lead sizga tegishli emas.'}, status=403)
    body = json_body(request)
    full_name = clean_string(body.get('full_name'))
    if not full_name:
        return ok({'full_name': ['Ism Familya kiritish shart.'], 'detail': 'Ism Familya kiritish shart.'}, status=400)
    old_name = lead.full_name
    lead.full_name = full_name
    lead.updated_at = timezone.now()
    lead.save(update_fields=['full_name', 'updated_at'])
    try:
        DataAuditLog.objects.create(
            actor=request.app_user,
            entity_type='lead',
            entity_id=lead.id,
            action='name_updated',
            old_data={'full_name': old_name},
            new_data={'full_name': full_name},
        )
    except Exception:
        pass
    return ok(get_one_lead(lead.id))


@require_auth('boss')
def boss_incoming_leads(request):
    """Boss uchun: o'zining operatorlari qo'lda qo'shgan (kiruvchi qo'ng'iroq) leadlar."""
    user = request.app_user
    qs = base_lead_qs().filter(boss=user, is_manual_entry=True)
    status = clean_string(request.GET.get('current_status'))
    if status:
        qs = qs.filter(current_status=status)
    if clean_string(request.GET.get('assigned_operator')):
        qs = qs.filter(assigned_operator_id=int(request.GET.get('assigned_operator')))
    qs = apply_lead_search(qs, request.GET.get('search'))
    return ok(serialize_leads(qs.order_by('-updated_at', '-id'), include_visit_decisions=True))


@require_auth('boss', 'filial_rahbari')
def boss_leads(request):
    user = request.app_user
    qs = base_lead_qs()
    if user.role == 'filial_rahbari':
        # Menenjerlar barcha Sotuv leadlarni ko‘radi.
        # Agar biror menenjer “Keldi” bosgan bo‘lsa, lead hamma menenjerlardan yopiladi.
        # Agar menenjer “Kelmadi” bosgan bo‘lsa, lead faqat o‘sha menenjer panelidan yopiladi.
        arrived_lead_ids = LeadVisitDecision.objects.filter(decision='arrived').values('lead_id')
        own_not_arrived_ids = LeadVisitDecision.objects.filter(decided_by=user, decision='not_arrived').values('lead_id')
        qs = qs.filter(current_status='sale').exclude(id__in=arrived_lead_ids).exclude(id__in=own_not_arrived_ids)
        qs = apply_lead_search(qs, request.GET.get('search')).order_by('-updated_at', '-id')
        sale_leads = list(qs)
        history_map = sale_history_map_for_leads([lead.id for lead in sale_leads])
        visible = [
            lead for lead in sale_leads
            if filial_can_see_lead(user, lead, history_map.get(lead.id, set()))
        ]
        return ok(serialize_leads(visible, include_visit_decisions=True))
    else:
        qs = qs.filter(boss=user)
        status = clean_string(request.GET.get('current_status'))
        if status:
            qs = qs.filter(current_status=status)
        if clean_string(request.GET.get('assigned_operator')):
            qs = qs.filter(assigned_operator_id=int(request.GET.get('assigned_operator')))
    qs = apply_lead_search(qs, request.GET.get('search'))
    return ok(serialize_leads(qs.order_by('-updated_at', '-id'), include_visit_decisions=True))


@csrf_exempt
@require_auth('admin')
def admin_leads(request, lead_id=None):
    if request.method == 'GET':
        qs = base_lead_qs()
        if request.GET.get('current_status'):
            qs = qs.filter(current_status=request.GET.get('current_status'))
        if request.GET.get('assigned_operator'):
            qs = qs.filter(assigned_operator_id=int(request.GET.get('assigned_operator')))
        if request.GET.get('boss'):
            qs = qs.filter(boss_id=int(request.GET.get('boss')))
        qs = apply_lead_search(qs, request.GET.get('search'))
        return ok(serialize_leads(qs.order_by('-updated_at', '-id'), include_visit_decisions=True))

    if request.method == 'DELETE':
        if not lead_id:
            body = json_body(request)
            raw_ids = body.get('ids') or body.get('lead_ids') or []
            if not isinstance(raw_ids, list):
                return ok({'detail': 'O‘chirish uchun leadlar ro‘yxati noto‘g‘ri yuborildi.'}, status=400)
            ids = []
            for item in raw_ids:
                try:
                    lead_pk = int(item)
                except (TypeError, ValueError):
                    continue
                if lead_pk > 0 and lead_pk not in ids:
                    ids.append(lead_pk)
            if not ids:
                return ok({'detail': 'O‘chirish uchun kamida bitta lead tanlang.'}, status=400)

            leads = list(Lead.objects.select_related('assigned_operator', 'boss').filter(id__in=ids))
            if not leads:
                return ok({'detail': 'Tanlangan leadlar topilmadi.'}, status=404)
            deleted_ids = [lead.id for lead in leads]
            audit_logs = []
            for lead in leads:
                audit_logs.append(DataAuditLog(
                    actor=request.app_user,
                    entity_type='lead',
                    entity_id=lead.id,
                    action='bulk_delete',
                    old_data={
                        'id': lead.id,
                        'full_name': lead.full_name,
                        'tsh': lead.tsh,
                        'phone1': lead.phone1,
                        'phone2': lead.phone2,
                        'phone3': lead.phone3,
                        'school': lead.school,
                        'grade': lead.grade,
                        'subject': lead.subject,
                        'ball': lead.ball,
                        'current_status': lead.current_status,
                        'operator': (lead.assigned_operator.full_name or lead.assigned_operator.username) if lead.assigned_operator_id and lead.assigned_operator else '',
                        'boss': (lead.boss.full_name or lead.boss.username) if lead.boss_id and lead.boss else '',
                    },
                    new_data={},
                ))
            with transaction.atomic():
                DataAuditLog.objects.bulk_create(audit_logs)
                Lead.objects.filter(id__in=deleted_ids).delete()
            return ok({'detail': f'{len(deleted_ids)} ta lead o‘chirildi.', 'deleted_count': len(deleted_ids), 'deleted_ids': deleted_ids})

        lead = Lead.objects.select_related('assigned_operator', 'boss').filter(id=lead_id).first()
        if not lead:
            return ok({'detail': 'Lead topilmadi.'}, status=404)
        old_data = {
            'id': lead.id,
            'full_name': lead.full_name,
            'tsh': lead.tsh,
            'phone1': lead.phone1,
            'phone2': lead.phone2,
            'phone3': lead.phone3,
            'school': lead.school,
            'grade': lead.grade,
            'subject': lead.subject,
            'ball': lead.ball,
            'current_status': lead.current_status,
            'operator': (lead.assigned_operator.full_name or lead.assigned_operator.username) if lead.assigned_operator_id and lead.assigned_operator else '',
            'boss': (lead.boss.full_name or lead.boss.username) if lead.boss_id and lead.boss else '',
        }
        lead.delete()
        DataAuditLog.objects.create(
            actor=request.app_user,
            entity_type='lead',
            entity_id=lead_id,
            action='delete',
            old_data=old_data,
            new_data={},
        )
        return ok({'detail': 'Lead o‘chirildi.', 'id': lead_id})

    return ok({'detail': 'Method not allowed'}, status=405)


@csrf_exempt
@require_auth('operator')
def change_status(request, lead_id):
    body = json_body(request)
    next_status = clean_string(body.get('current_status') or body.get('status'))
    note = clean_string(body.get('note'))
    if request.method not in ('PATCH', 'PUT', 'POST'):
        return ok({'detail': 'Method not allowed'}, status=405)
    if not is_valid_status(next_status):
        return ok({'current_status': ['Status noto‘g‘ri.'], 'detail': 'Status noto‘g‘ri.'}, status=400)
    if next_status in NOTE_REQUIRED_STATUSES and not note:
        return ok({'note': ['Bu status uchun izoh kiritish kerak.'], 'detail': 'Bu status uchun izoh kiritish kerak.'}, status=400)

    lead = Lead.objects.filter(id=lead_id, assigned_operator=request.app_user).first()
    if not lead:
        return ok({'detail': 'Bu lead sizga tegishli emas.'}, status=403)

    old_status = lead.current_status
    selected_branch = ''
    if next_status == 'sale':
        selected_branch = normalize_branch_label(body.get('selected_branch') or body.get('branch_name') or body.get('filial') or '')
        if not selected_branch:
            return ok({'selected_branch': ['Sotuv qaysi filial uchun bo‘lganini tanlang.'], 'detail': 'Sotuv qaysi filial uchun bo‘lganini tanlang.'}, status=400)
        if selected_branch not in BRANCH_CHOICES:
            return ok({'selected_branch': ['Tanlangan filial topilmadi.'], 'detail': 'Tanlangan filial topilmadi.'}, status=400)

    try:
        with transaction.atomic():
            lead.current_status = next_status
            lead.updated_at = timezone.now()
            update_fields = ['current_status', 'updated_at']
            if next_status == 'sale' and selected_branch:
                lead.branch_name = selected_branch
                update_fields.append('branch_name')
            lead.save(update_fields=update_fields)

            history_note = note
            if next_status == 'sale' and selected_branch:
                history_note = f"{note} | Filial: {selected_branch}" if note else f"Filial: {selected_branch}"
            # note ustuni 500 belgidan oshib ketsa baza xato bermasligi uchun qisqartiramiz
            history_note = clean_string(history_note)[:500]

            LeadStatusHistory.objects.create(
                lead=lead,
                old_status=old_status,
                new_status=next_status,
                changed_by=request.app_user,
                note=history_note,
            )
            try:
                DataAuditLog.objects.create(
                    actor=request.app_user,
                    entity_type='lead',
                    entity_id=lead.id,
                    action='status_changed',
                    old_data={'current_status': old_status},
                    new_data={'current_status': next_status, 'note': history_note, 'selected_branch': selected_branch},
                )
            except Exception:
                # Audit log status almashtirishni to‘xtatib qo‘ymasligi kerak.
                pass

            if next_status in ('sale', 'advice') and old_status != next_status:
                title = 'Yangi SOTUV lead' if next_status == 'sale' else 'Yangi MASLAHAT lead'
                emoji = '💰' if next_status == 'sale' else '💬'
                msg = (
                    f'{emoji} <b>{title}</b>\n'
                    + tg_line('Vaqt', tg_now_text())
                    + tg_line('Operator', tg_user_name(request.app_user))
                    + tg_line('Oldingi status', status_text(old_status))
                    + tg_line('Yangi status', status_text(next_status))
                    + tg_line('Izoh', note)
                    + (tg_line('Filial', selected_branch) if selected_branch else '')
                    + '\n' + tg_lead_info(lead)
                )
                try:
                    notify_telegram_after_commit(msg)
                except Exception:
                    pass
    except Exception as exc:
        return ok({'detail': f'Statusni saqlashda backend xatoligi: {exc}'}, status=400)

    return ok(get_one_lead(lead.id))



@require_auth('operator')
def reminders(request):
    qs = base_lead_qs().filter(assigned_operator=request.app_user, reminder_at__isnull=False, reminder_last_notified_at__isnull=True).order_by('reminder_at')
    return ok(serialize_leads(qs, include_visit_decisions=False))


@csrf_exempt
@require_auth('operator')
def set_reminder(request, lead_id):
    body = json_body(request)
    lead = Lead.objects.filter(id=lead_id, assigned_operator=request.app_user).first()
    if not lead:
        return ok({'detail': 'Bu lead sizga tegishli emas.'}, status=403)
    old = {'reminder_at': lead.reminder_at.isoformat() if lead.reminder_at else None, 'reminder_note': lead.reminder_note, 'reminder_last_notified_at': lead.reminder_last_notified_at.isoformat() if lead.reminder_last_notified_at else None}
    if 'reminder_at' in body:
        if body.get('reminder_at'):
            lead.reminder_at = parse_datetime_value(body.get('reminder_at'))
            lead.reminder_last_notified_at = None
        else:
            lead.reminder_at = None
            lead.reminder_note = ''
            lead.reminder_last_notified_at = None
    if 'reminder_note' in body:
        lead.reminder_note = clean_string(body.get('reminder_note'))
    if body.get('mark_as_notified') and lead.reminder_at:
        lead.reminder_last_notified_at = timezone.now()
    lead.updated_at = timezone.now()
    lead.save()
    new = {'reminder_at': lead.reminder_at.isoformat() if lead.reminder_at else None, 'reminder_note': lead.reminder_note, 'reminder_last_notified_at': lead.reminder_last_notified_at.isoformat() if lead.reminder_last_notified_at else None}
    DataAuditLog.objects.create(actor=request.app_user, entity_type='lead', entity_id=lead.id, action='reminder_updated', old_data=old, new_data=new)
    return ok(get_one_lead(lead.id))


@csrf_exempt
@require_http_methods(['POST'])
def public_online_leads(request):
    body = json_body(request)
    tsh = clean_string(body.get('tsh') or body.get('t_sh'))
    school = clean_string(body.get('school') or body.get('maktab'))
    grade = clean_string(body.get('grade') or body.get('sinf'))
    full_name = clean_string(body.get('full_name') or body.get('fio'))
    subject = clean_string(body.get('subject') or body.get('interest_subject') or body.get('fan'))
    phone1 = clean_string(body.get('phone1') or body.get('tel1'))
    phone2 = clean_string(body.get('phone2') or body.get('tel2'))
    phone3 = clean_string(body.get('phone3') or body.get('tel3'))
    age = int(body.get('age') or 0)
    region = clean_string(body.get('region'))

    boss = AppUser.objects.filter(role='boss', is_active=True).order_by('id').first()
    item = OnlineLead.objects.create(
        full_name=full_name, tsh=tsh, school=school, grade=grade, subject=subject,
        age=age, phone1=phone1, phone2=phone2, phone3=phone3,
        interest_subject=subject, region=region, assigned_boss=boss, source_payload=body
    )
    msg = (
        '🆕 <b>Yangi online lead keldi</b>\n'
        + tg_line('Vaqt', tg_now_text())
        + tg_line('T/SH', tsh)
        + tg_line('Maktab', school)
        + tg_line('Sinf', grade)
        + tg_line('F.I.O', full_name)
        + tg_line('Fan', subject)
        + tg_line('Tel 1', phone1)
        + tg_line('Tel 2', phone2)
        + tg_line('Tel 3', phone3)
        + tg_line('Boss', tg_user_name(boss))
    )
    notify_telegram_after_commit(msg)
    return ok(online_lead_to_dict(item), status=201)


@require_auth('boss')
def boss_online_leads(request):
    qs = OnlineLead.objects.select_related('assigned_operator').filter(assigned_operator__isnull=True, assigned_boss=request.app_user)
    search = clean_string(request.GET.get('search'))
    if search:
        qs = qs.filter(Q(full_name__icontains=search) | Q(phone1__icontains=search) | Q(phone2__icontains=search) | Q(phone3__icontains=search) | Q(tsh__icontains=search) | Q(school__icontains=search) | Q(grade__icontains=search) | Q(subject__icontains=search) | Q(interest_subject__icontains=search))
    return ok([online_lead_to_dict(x) for x in qs.order_by('-submitted_at')])


@require_auth('operator')
def operator_online_leads(request):
    qs = OnlineLead.objects.select_related('assigned_operator').filter(assigned_operator=request.app_user).order_by('-assigned_at', '-submitted_at')
    return ok([online_lead_to_dict(x) for x in qs])


@csrf_exempt
@require_auth('boss')
def assign_online_lead(request, online_id):
    body = json_body(request)
    operator_id = int(body.get('operator_id') or 0)
    operator = AppUser.objects.filter(id=operator_id, role='operator', boss=request.app_user, is_active=True).first()
    if not operator:
        return ok({'detail': 'Tanlangan operator topilmadi.'}, status=403)
    online = OnlineLead.objects.filter(id=online_id, assigned_operator__isnull=True, assigned_boss=request.app_user).first()
    if not online:
        return ok({'detail': 'Lead topilmadi yoki allaqachon biriktirilgan.'}, status=404)

    duplicates = find_existing_general_leads(online.full_name, online.phone1, online.phone2, online.phone3)
    duplicate_of_lead = duplicates[0] if duplicates else None

    with transaction.atomic():
        lead = Lead.objects.create(
            full_name=online.full_name, tsh=getattr(online, 'tsh', '') or '',
            phone1=online.phone1, phone2=online.phone2, phone3=online.phone3,
            subject=(getattr(online, 'subject', '') or online.interest_subject or ''),
            school=getattr(online, 'school', '') or '', grade=(getattr(online, 'grade', '') or (f'{online.age} yosh' if online.age else '')),
            branch_name=online.region or '',
            assigned_operator=operator, boss=request.app_user, uploaded_by=request.app_user, current_status='new',
            is_duplicate=bool(duplicate_of_lead), duplicate_of_lead=duplicate_of_lead
        )
        online.assigned_operator = operator
        online.created_lead = lead
        online.assigned_at = timezone.now()
        online.save()
        DataAuditLog.objects.create(actor=request.app_user, entity_type='online_lead', entity_id=online.id, action='assigned', new_data={'operator_id': operator_id, 'lead_id': lead.id})
        msg = (
            '👤 <b>Online lead operatorga biriktirildi</b>\n'
            + tg_line('Vaqt', tg_now_text())
            + tg_line('Boss', tg_user_name(request.app_user))
            + tg_line('Operator', tg_user_name(operator))
            + '\n' + tg_lead_info(lead)
        )
        notify_telegram_after_commit(msg)
    return ok(online_lead_to_dict(online))


@csrf_exempt
@require_auth('boss')
def bulk_assign_online_leads(request):
    body = json_body(request)
    operator_id = int(body.get('operator_id') or 0)
    operator = AppUser.objects.filter(id=operator_id, role='operator', boss=request.app_user, is_active=True).first()
    if not operator:
        return ok({'detail': 'Tanlangan operator topilmadi.'}, status=403)
    items = list(OnlineLead.objects.filter(assigned_operator__isnull=True, assigned_boss=request.app_user).order_by('submitted_at'))
    duplicate_index = build_duplicate_lead_index()
    count = 0
    skipped_duplicates = []
    with transaction.atomic():
        for online in items:
            duplicates = find_duplicate_leads_in_index(duplicate_index, online.full_name, online.phone1, online.phone2, online.phone3)
            duplicate_of_lead = duplicates[0] if duplicates else None
            lead = Lead.objects.create(full_name=online.full_name, tsh=getattr(online, 'tsh', '') or '', phone1=online.phone1, phone2=online.phone2, phone3=online.phone3, subject=(getattr(online, 'subject', '') or online.interest_subject or ''), school=getattr(online, 'school', '') or '', grade=(getattr(online, 'grade', '') or (f'{online.age} yosh' if online.age else '')), branch_name=online.region or '', assigned_operator=operator, boss=request.app_user, uploaded_by=request.app_user, current_status='new', is_duplicate=bool(duplicate_of_lead), duplicate_of_lead=duplicate_of_lead)
            add_lead_to_duplicate_index(duplicate_index, lead)
            online.assigned_operator = operator
            online.created_lead = lead
            online.assigned_at = timezone.now()
            online.save()
            count += 1
    detail = f'{count} ta online lead biriktirildi.'
    return ok({'assigned': count, 'skipped_duplicates': 0, 'duplicate_items': [], 'detail': detail})


@csrf_exempt
@require_auth('filial_rahbari')
def visit_decision(request, lead_id):
    if request.method not in ('POST', 'PATCH', 'PUT'):
        return ok({'detail': 'Method not allowed'}, status=405)

    body = json_body(request)
    decision = clean_string(body.get('decision'))
    if decision not in ('arrived', 'not_arrived'):
        return ok({'decision': ['Qaror noto‘g‘ri.'], 'detail': 'Qaror noto‘g‘ri.'}, status=400)

    lead = Lead.objects.select_related('assigned_operator', 'boss').filter(
        id=lead_id,
        current_status='sale',
    ).first()
    if not lead:
        return ok({'detail': 'Lead topilmadi yoki u Sotuv holatida emas.'}, status=404)
    if not filial_can_see_lead(request.app_user, lead):
        return ok({'detail': 'Bu lead sizga biriktirilgan filialga tegishli emas.'}, status=403)

    changed = False
    old_decision = ''
    with transaction.atomic():
        # Bir leadga bir nechta menenjer ilgari Kelmadi bosgan bo‘lishi mumkin.
        # Keldi yakuniy holat bo‘lgani uchun u barcha menenjerlar uchun ustun turadi.
        locked_items = list(
            LeadVisitDecision.objects.select_for_update()
            .filter(lead=lead)
            .order_by('-updated_at', '-id')
        )
        arrived_item = next((row for row in locked_items if row.decision == 'arrived'), None)
        if arrived_item is not None:
            if decision == 'not_arrived':
                return ok({'detail': 'Keldi bosilgandan keyin Kelmadi qilib o‘zgartirib bo‘lmaydi.'}, status=409)
            item = arrived_item
        else:
            # Eski menenjerning yozuvini yangi menenjerga ko‘chirmaymiz.
            # Har bir menenjer o‘z qarorini saqlaydi. Bu unique constraint bilan
            # to‘qnashuvni bartaraf qiladi va keyin yaratilgan menenjerga ham
            # Kelmadi/Keldi bosish imkonini beradi.
            item = next((row for row in locked_items if row.decided_by_id == request.app_user.id), None)
            if item is None:
                try:
                    # Ichki atomic savepoint IntegrityError bo‘lsa tashqi
                    # tranzaksiyani "broken" holatga tushirmaydi.
                    with transaction.atomic():
                        item = LeadVisitDecision.objects.create(
                            lead=lead,
                            decided_by=request.app_user,
                            decision=decision,
                        )
                    changed = True
                except IntegrityError:
                    # Ikki marta tez bosilganda yoki parallel so‘rov kelganda
                    # mavjud yozuvni qayta olib xavfsiz yangilaymiz.
                    item = LeadVisitDecision.objects.select_for_update().get(
                        lead=lead,
                        decided_by=request.app_user,
                    )
            old_decision = item.decision if not changed else ''
            if item.decision != decision:
                item.decision = decision
                item.updated_at = timezone.now()
                item.save(update_fields=['decision', 'updated_at'])
                changed = True

    # Asosiy belgi allaqachon saqlandi. Tarix/audit jadvalidagi vaqtinchalik
    # muammo asosiy Keldi/Kelmadi amalini bekor qilmasligi kerak.
    if changed:
        try:
            LeadVisitDecisionHistory.objects.create(
                lead=lead,
                decided_by=request.app_user,
                old_decision=old_decision,
                new_decision=decision,
            )
        except Exception:
            pass
        try:
            DataAuditLog.objects.create(
                actor=request.app_user,
                entity_type='lead_visit_decision',
                entity_id=item.id,
                action='upserted',
                old_data={'decision': old_decision},
                new_data={'decision': decision},
            )
        except Exception:
            pass

        decision_label = 'Keldi' if decision == 'arrived' else 'Kelmadi'
        emoji = '✅' if decision == 'arrived' else '❌'
        msg = (
            f'{emoji} <b>Menenjer {decision_label} bosdi</b>\n'
            + tg_line('Vaqt', tg_now_text())
            + tg_line('Menenjer', tg_user_name(request.app_user))
            + tg_line('Filial', request.app_user.branch_name)
            + tg_line('Qaror', decision_label)
            + tg_line('Operator', tg_user_name(lead.assigned_operator))
            + '\n' + tg_lead_info(lead)
        )
        notify_telegram_after_commit(msg)

    return ok(visit_decision_to_dict(item))


@csrf_exempt
@require_auth('filial_rahbari')
def mark_payment_status(request, lead_id):
    """Keldi deb belgilangan lead uchun to‘lov holatini bitta xavfsiz endpointda saqlaydi.

    Frontend ayrim eski ekranlarda lead ID o‘rniga LeadVisitDecision ID yuborgan
    bo‘lishi mumkin. Shu sabab ``lead_id`` avval lead sifatida, topilmasa nazorat
    yozuvi ID sifatida ham tekshiriladi.
    """
    if request.method not in ('POST', 'PATCH', 'PUT'):
        return ok({'detail': 'Method not allowed'}, status=405)

    body = json_body(request)
    status_value = clean_string(body.get('status') or body.get('payment_status')).lower()
    aliases = {
        'paid': 'paid',
        'done': 'paid',
        'payment_done': 'paid',
        'unpaid': 'unpaid',
        'not_done': 'unpaid',
        'payment_not_done': 'unpaid',
        'left': 'left_without_payment',
        'left_without_payment': 'left_without_payment',
    }
    payment_status = aliases.get(status_value)
    if not payment_status:
        return ok({'detail': 'To‘lov holati noto‘g‘ri yuborildi.'}, status=400)

    return _save_visit_payment_status(request, lead_id, payment_status)


def _resolve_payment_lead(identifier):
    """Lead ID yoki nazorat yozuvi ID orqali sotuv leadini topadi."""
    lead = Lead.objects.select_related('assigned_operator', 'boss').filter(
        id=identifier,
        current_status='sale',
    ).first()
    if lead:
        return lead

    decision = LeadVisitDecision.objects.select_related(
        'lead', 'lead__assigned_operator', 'lead__boss'
    ).filter(id=identifier).first()
    if decision and decision.lead and decision.lead.current_status == 'sale':
        return decision.lead
    return None


def _payment_response_payload(item):
    """Belgi saqlangandan keyin serializerdagi ikkilamchi xato 500 qaytarmasin."""
    try:
        payload = visit_decision_to_dict(item)
    except Exception:
        logger.exception('To‘lov belgisi saqlandi, ammo javob serializerida xato yuz berdi.')
        payload = {
            'id': item.id,
            'lead': item.lead_id,
            'lead_id': item.lead_id,
            'decision': item.decision,
            'payment_status': visit_payment_status(item),
            'payment_done': bool(item.payment_done),
            'payment_not_done': bool(item.payment_not_done),
            'left_without_payment': bool(item.left_without_payment),
            'payment_done_at': item.payment_done_at,
            'payment_not_done_at': item.payment_not_done_at,
            'left_without_payment_at': item.left_without_payment_at,
            'payment_status_at': visit_payment_status_at(item),
            'payment_status_by_name': visit_payment_status_by_name(item),
            'updated_at': item.updated_at,
        }
    payload['saved'] = True
    return payload


def _save_visit_payment_status(request, lead_identifier, payment_status):
    lead = _resolve_payment_lead(lead_identifier)
    if not lead:
        return ok({'detail': 'Lead topilmadi yoki u Sotuv holatida emas.'}, status=404)
    if not filial_can_see_lead(request.app_user, lead):
        return ok({'detail': 'Bu lead sizga biriktirilgan filialga tegishli emas.'}, status=403)

    changed = False
    old_status = 'pending'
    try:
        with transaction.atomic():
            item = manager_existing_visit_decision(request.app_user, lead, for_update=True)
            if not item:
                return ok({'detail': 'Avval Keldi deb belgilang.'}, status=400)
            if item.decision != 'arrived':
                return ok({'detail': 'To‘lov holatini faqat Keldi belgilangan leadga qo‘yish mumkin.'}, status=400)

            old_status = visit_payment_status(item)
            if old_status != payment_status:
                now = timezone.now()
                item.payment_done = payment_status == 'paid'
                item.payment_done_at = now if payment_status == 'paid' else None
                item.payment_done_by = request.app_user if payment_status == 'paid' else None
                item.payment_not_done = payment_status == 'unpaid'
                item.payment_not_done_at = now if payment_status == 'unpaid' else None
                item.payment_not_done_by = request.app_user if payment_status == 'unpaid' else None
                item.left_without_payment = payment_status == 'left_without_payment'
                item.left_without_payment_at = now if payment_status == 'left_without_payment' else None
                item.left_without_payment_by = request.app_user if payment_status == 'left_without_payment' else None
                item.updated_at = now
                item.save(update_fields=[
                    'payment_done', 'payment_done_at', 'payment_done_by',
                    'payment_not_done', 'payment_not_done_at', 'payment_not_done_by',
                    'left_without_payment', 'left_without_payment_at', 'left_without_payment_by',
                    'updated_at',
                ])
                changed = True
    except Exception:
        request_id = getattr(request, 'request_id', '')
        logger.exception(
            'To‘lov holatini saqlashda xato. request_id=%s lead=%s status=%s manager=%s',
            request_id,
            getattr(lead, 'id', lead_identifier),
            payment_status,
            getattr(request.app_user, 'id', None),
        )
        return ok({
            'detail': 'To‘lov holatini bazaga saqlashda server xatosi yuz berdi.',
            'request_id': request_id,
        }, status=500)

    if changed:
        action_map = {
            'paid': 'payment_paid',
            'unpaid': 'payment_unpaid',
            'left_without_payment': 'left_without_payment',
        }
        try:
            DataAuditLog.objects.create(
                actor=request.app_user,
                entity_type='lead_visit_payment',
                entity_id=item.id,
                action=action_map[payment_status],
                old_data={'payment_status': old_status},
                new_data={'payment_status': payment_status},
            )
        except Exception:
            logger.exception('To‘lov audit yozuvini saqlashda xato.')

        label_map = {
            'paid': ('💳', 'To‘lov qildi'),
            'unpaid': ('❌', 'To‘lov qilmadi'),
            'left_without_payment': ('🚶', 'Keldi, to‘lov qilmasdan ketdi'),
        }
        emoji, label = label_map[payment_status]
        msg = (
            f'{emoji} <b>{label}</b>\n'
            + tg_line('Vaqt', tg_now_text())
            + tg_line('Menenjer', tg_user_name(request.app_user))
            + tg_line('Filial', request.app_user.branch_name)
            + tg_line('Operator', tg_user_name(lead.assigned_operator))
            + '\n' + tg_lead_info(lead)
        )
        try:
            notify_telegram_after_commit(msg)
        except Exception:
            logger.exception('To‘lov Telegram xabarini navbatga qo‘yishda xato.')

    return ok(_payment_response_payload(item))


@csrf_exempt
@require_auth('filial_rahbari')
def mark_visit_payment(request, lead_id):
    """Eski endpoint: “To‘lov qildi”."""
    if request.method not in ('POST', 'PATCH', 'PUT'):
        return ok({'detail': 'Method not allowed'}, status=405)
    return _save_visit_payment_status(request, lead_id, 'paid')


@csrf_exempt
@require_auth('filial_rahbari')
def mark_payment_not_done(request, lead_id):
    """Eski endpoint: “To‘lov qilmadi”."""
    if request.method not in ('POST', 'PATCH', 'PUT'):
        return ok({'detail': 'Method not allowed'}, status=405)
    return _save_visit_payment_status(request, lead_id, 'unpaid')


@csrf_exempt
@require_auth('filial_rahbari')
def mark_left_without_payment(request, lead_id):
    """Eski endpoint: “Keldi, to‘lov qilmasdan ketdi”."""
    if request.method not in ('POST', 'PATCH', 'PUT'):
        return ok({'detail': 'Method not allowed'}, status=405)
    return _save_visit_payment_status(request, lead_id, 'left_without_payment')


@require_auth('boss', 'filial_rahbari')
def lead_visit_decisions(request):
    user = request.app_user
    if user.role == 'filial_rahbari':
        # Eski akkaunt yaratgan qarorlar ham ayni filialga biriktirilgan yangi
        # menenjerga ko‘rinadi. Huquq decided_by ID emas, lead filiali orqali.
        items = manager_visible_visit_decisions(user)
        return ok([visit_decision_to_dict(x) for x in items])

    operator_ids = list(AppUser.objects.filter(role='operator', boss=user).values_list('id', flat=True))
    qs = LeadVisitDecision.objects.select_related(
        'lead', 'lead__assigned_operator', 'lead__boss', 'decided_by',
        'payment_done_by', 'payment_not_done_by', 'left_without_payment_by',
    ).filter(lead__assigned_operator_id__in=operator_ids).order_by('-updated_at')
    return ok([visit_decision_to_dict(x) for x in qs])


@require_auth('operator')
def operator_visit_decisions(request):
    user = request.app_user
    qs = LeadVisitDecision.objects.select_related(
        'lead', 'lead__assigned_operator', 'lead__boss', 'decided_by', 'payment_done_by', 'payment_not_done_by', 'left_without_payment_by'
    ).filter(lead__assigned_operator=user).order_by('-updated_at')
    # Operator faqat o‘ziga biriktirilgan filiallar menenjerlari bosgan Keldi/Kelmadi natijalarini ko‘radi.
    visible = [item for item in qs if item.decided_by and users_branch_overlap(user, item.decided_by)]
    decision = clean_string(request.GET.get('decision'))
    payment = clean_string(request.GET.get('payment'))
    if decision in ('arrived', 'not_arrived'):
        visible = [item for item in visible if item.decision == decision]
    if payment in ('done', 'paid'):
        visible = [item for item in visible if visit_payment_status(item) == 'paid']
    elif payment in ('not_done', 'unpaid'):
        visible = [item for item in visible if visit_payment_status(item) == 'unpaid']
    elif payment in ('left_without_payment', 'left'):
        visible = [item for item in visible if visit_payment_status(item) == 'left_without_payment']
    elif payment == 'pending':
        visible = [item for item in visible if visit_payment_status(item) == 'pending']
    return ok([visit_decision_to_dict(x) for x in visible])


def autosize(ws):
    for col in ws.columns:
        max_len = 10
        letter = col[0].column_letter
        for cell in col:
            max_len = max(max_len, len(str(cell.value or '')))
        ws.column_dimensions[letter].width = min(max_len + 2, 45)


def excel_response(wb, filename):
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f"attachment; filename={filename}"
    return resp


@require_auth('boss')
def import_template(request):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Leadlar'
    ws.append(['№', 'T/SH', 'Maktab', 'Sinf', 'F.I.O', 'Fan', 'Ball', 'tel1', 'tel2', 'tel3', 'Filial'])
    ws.append([1, 'T001', '25-maktab', '8-sinf', 'Aliyev Ali', 'English', '85', '+998901234567', '', '', 'Niyozbosh'])
    autosize(ws)
    return excel_response(wb, 'lead_import_shablon.xlsx')


def excel_cell_text(value):
    if value is None:
        return ''
    # Excel telefonni son qilib o‘qisa 998901234567.0 ko‘rinishida kelishi mumkin.
    # Uni stringga aylantirishda oxiriga ortiqcha 0 qo‘shilib qolmasligi uchun integer qilib olamiz.
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_excel_rows(file_obj):
    wb = load_workbook(file_obj, data_only=True)
    ws = wb.worksheets[0]
    headers = []
    for cell in ws[1]:
        headers.append(normalize_excel_column(cell.value))
    rows = []
    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        item = {}
        has = False
        for key, value in zip(headers, row):
            if not key:
                continue
            text = excel_cell_text(value)
            item[key] = text
            if text:
                has = True
        if has:
            rows.append((index, item))
    return headers, rows


@csrf_exempt
@require_auth('boss')
def import_leads(request):
    operator_value = clean_string(request.POST.get('operator_id'))
    file = request.FILES.get('file')
    if not file:
        return ok({'detail': 'Excel fayl tanlang.'}, status=400)
    if not file.name.lower().endswith('.xlsx'):
        return ok({'detail': 'Faqat .xlsx Excel fayl yuklash mumkin.'}, status=400)
    headers, rows = read_excel_rows(file)
    required = ['T/SH', 'Maktab', 'Sinf', 'F.I.O', 'Fan', 'Ball', 'tel1', 'tel2', 'tel3']
    missing = [c for c in required if c not in headers]
    if missing:
        return ok({'detail': 'Excel formati noto‘g‘ri.', 'missing_columns': missing, 'received_columns': headers}, status=400)
    operators = []
    operator_mode = 'single'
    assigned_operator = None
    if operator_value == 'all':
        operators = list(AppUser.objects.filter(role='operator', boss=request.app_user, is_active=True).order_by('id'))
        operator_mode = 'all'
        if not operators:
            return ok({'detail': 'Operatorlar topilmadi.'}, status=400)
    else:
        assigned_operator = AppUser.objects.filter(id=int(operator_value or 0), role='operator', boss=request.app_user, is_active=True).first()
        if not assigned_operator:
            return ok({'detail': 'Operator topilmadi.'}, status=400)
    import_record = ExcelImport.objects.create(uploaded_by=request.app_user, assigned_operator=assigned_operator, filename=file.name, total_rows=len(rows))
    duplicate_index = build_duplicate_lead_index()
    success = duplicate = failed = 0
    duplicate_items = []
    op_index = 0
    for source_row_number, raw in rows:
        normalized = {
            'tsh': clean_string(raw.get('T/SH')),
            'school': clean_string(raw.get('Maktab')),
            'grade': clean_string(raw.get('Sinf')),
            'full_name': clean_string(raw.get('F.I.O')),
            'subject': clean_string(raw.get('Fan')),
            'ball': clean_string(raw.get('Ball')),
            'phone1': normalize_phone(raw.get('tel1')),
            'phone2': normalize_phone(raw.get('tel2')),
            'phone3': normalize_phone(raw.get('tel3')),
            'branch_name': clean_string(raw.get('Filial')),
        }
        try:
            # F.I.O yoki telefon raqam majburiy emas.
            # Excel qatorida kamida bitta ma'lumot bo'lsa, lead saqlanadi.
            operator = operators[op_index % len(operators)] if operator_mode == 'all' else assigned_operator
            if operator_mode == 'all':
                op_index += 1

            duplicates = find_duplicate_leads_in_index(
                duplicate_index,
                normalized.get('full_name'),
                normalized.get('phone1'),
                normalized.get('phone2'),
                normalized.get('phone3'),
            )
            duplicate_of_lead = duplicates[0] if duplicates else None
            if duplicates:
                duplicate += 1
                duplicate_items.append({
                    'source_row_number': source_row_number,
                    'full_name': normalized.get('full_name') or '',
                    'phone1': normalized.get('phone1') or '',
                    'phone2': normalized.get('phone2') or '',
                    'phone3': normalized.get('phone3') or '',
                    'duplicate_count': len(duplicates),
                    'duplicate_leads': duplicate_leads_payload(duplicates),
                    'detail': 'Bu lead bazada bor edi, lekin yangi operatorga ham biriktirildi.',
                })

            lead = Lead.objects.create(
                **normalized,
                assigned_operator=operator,
                boss=request.app_user,
                uploaded_by=request.app_user,
                current_status='new',
                source_row_number=source_row_number,
                is_duplicate=bool(duplicate_of_lead),
                duplicate_of_lead=duplicate_of_lead,
            )
            add_lead_to_duplicate_index(duplicate_index, lead)
            ExcelImportRow.objects.create(
                import_record=import_record,
                source_row_number=source_row_number,
                raw_data=raw,
                normalized_data=normalized,
                status='saved',
                lead=lead,
                duplicate_of_lead=duplicate_of_lead,
            )
            success += 1
        except Exception as e:
            failed += 1
            ExcelImportRow.objects.create(import_record=import_record, source_row_number=source_row_number, raw_data=raw, normalized_data=normalized, status='failed', error_message=str(e))
    import_record.success_rows = success
    import_record.duplicate_rows = duplicate
    import_record.failed_rows = failed
    import_record.save()
    if operator_mode == 'all':
        operator_text = f'Barcha operatorlar ({len(operators)} ta)'
    else:
        operator_text = tg_user_name(assigned_operator)
    msg = (
        '📥 <b>Excel import qilindi</b>\n'
        + tg_line('Vaqt', tg_now_text())
        + tg_line('Boss', tg_user_name(request.app_user))
        + tg_line('Fayl', file.name)
        + tg_line('Operator', operator_text)
        + tg_line('Jami qator', len(rows))
        + tg_line('Saqlangan', success)
        + tg_line('Dublikat', duplicate)
        + tg_line('Xato', failed)
    )
    notify_telegram_after_commit(msg)
    return ok({
        'total_rows': len(rows),
        'success_rows': success,
        'duplicate_rows': duplicate,
        'failed_rows': failed,
        'operator_mode': operator_mode,
        'duplicate_items': duplicate_items[:50],
    })


@require_auth('admin')
def admin_statistics(request):
    data = {
        'bosses': AppUser.objects.filter(role='boss', is_active=True).count(),
        'filial_rahbarlari': AppUser.objects.filter(role='filial_rahbari', is_active=True).count(),
        'operators': AppUser.objects.filter(role='operator', is_active=True).count(),
        'new': Lead.objects.filter(current_status='new').count(),
    }
    for st in STATUS_ORDER:
        data[st] = Lead.objects.filter(current_status=st).count()
    return ok(data)


@require_auth('director', 'director_deputy', 'admin')
def director_statistics(request):
    """Bosh direktor / o'rinbosari uchun butun tizim bo'yicha to'liq nazorat statistikasi:
    har bir boss va uning operatorlari kesimida tanlangan davr (kun/oy/oraliq) bo'yicha
    sotuv, atkaz, maslahat va boshqa holatlar. Filtr berilmasa joriy oy va bugungi kun ishlatiladi."""
    bosses = list(AppUser.objects.filter(role='boss', is_active=True).order_by('full_name', 'username'))
    boss_ids = [b.id for b in bosses]
    ops = list(AppUser.objects.filter(role='operator', boss_id__in=boss_ids, is_active=True).order_by('full_name', 'username'))
    op_ids = [op.id for op in ops]

    date_info = parse_date_range(request.GET)
    period_start_dt = date_info['start_dt']
    period_end_dt = date_info['end_dt']
    is_single_day = date_info['start_date'] == date_info['end_date']

    period_assigned = Lead.objects.filter(assigned_operator_id__in=op_ids, created_at__range=(period_start_dt, period_end_dt)) if op_ids else Lead.objects.none()
    period_hist = list(LeadStatusHistory.objects.filter(changed_by_id__in=op_ids, changed_at__range=(period_start_dt, period_end_dt))) if op_ids else []

    period_assigned_counts = {}
    for lead in period_assigned.values('assigned_operator_id').annotate(cnt=Count('id')):
        period_assigned_counts[lead['assigned_operator_id']] = lead['cnt']

    operator_rows = []
    for op in ops:
        r = {
            'operator_id': op.id,
            'operator_name': op.full_name or op.username,
            'boss_id': op.boss_id,
            'boss_name': (op.boss.full_name or op.boss.username) if op.boss_id and op.boss else '-',
            'total': period_assigned_counts.get(op.id, 0),
            **empty_status_counts(),
            'actions_period': 0,
            'touched_period': 0,
            'period_sale': 0,
        }
        touched_ids = set()
        for h in period_hist:
            if h.changed_by_id == op.id:
                r['actions_period'] += 1
                touched_ids.add(h.lead_id)
                if h.new_status in r:
                    r[h.new_status] += 1
                if h.new_status == 'sale':
                    r['period_sale'] += 1
        r['touched_period'] = len(touched_ids)
        r['conversion'] = calc_conversion(r['sale'], r['total'])
        operator_rows.append(r)

    boss_rows = []
    for b in bosses:
        boss_op_rows = [r for r in operator_rows if r['boss_id'] == b.id]
        boss_row = {
            'boss_id': b.id,
            'boss_name': b.full_name or b.username,
            'operator_count': len(boss_op_rows),
            'total': sum(r['total'] for r in boss_op_rows),
            'period_sale': sum(r['period_sale'] for r in boss_op_rows),
            'actions_period': sum(r['actions_period'] for r in boss_op_rows),
            'touched_period': sum(r['touched_period'] for r in boss_op_rows),
        }
        for st in STATUS_ORDER:
            boss_row[st] = sum(r.get(st, 0) for r in boss_op_rows)
        boss_row['conversion'] = calc_conversion(boss_row['sale'], boss_row['total'])
        boss_rows.append(boss_row)

    summary = {
        'bosses': len(bosses),
        'operators': len(ops),
        'total': sum(r['total'] for r in operator_rows),
        'period_sale': sum(r['period_sale'] for r in operator_rows),
        'actions_period': sum(r['actions_period'] for r in operator_rows),
        'touched_period': sum(r['touched_period'] for r in operator_rows),
    }
    for st in STATUS_ORDER:
        summary[st] = sum(r.get(st, 0) for r in operator_rows)
    summary['conversion'] = calc_conversion(summary['sale'], summary['total'])

    top_operators_period = sorted(operator_rows, key=lambda r: r['period_sale'], reverse=True)[:10]

    return ok({
        'start_date': date_info['start_date'],
        'end_date': date_info['end_date'],
        'is_single_day': is_single_day,
        'summary': summary,
        'boss_rows': boss_rows,
        'operator_rows': operator_rows,
        'top_operators_period': top_operators_period,
    })


@require_auth('boss')
def boss_statistics(request):
    ops = AppUser.objects.filter(role='operator', boss=request.app_user, is_active=True).order_by('full_name', 'username')
    op_ids = [op.id for op in ops]
    date_info = parse_date_range({'month': timezone.localdate().strftime('%Y-%m')})
    today, start, end = day_range()
    month_assigned = Lead.objects.filter(assigned_operator_id__in=op_ids, created_at__range=(date_info['start_dt'], date_info['end_dt'])) if op_ids else Lead.objects.none()
    month_hist = list(LeadStatusHistory.objects.filter(changed_by_id__in=op_ids, changed_at__range=(date_info['start_dt'], date_info['end_dt']))) if op_ids else []
    today_hist = list(LeadStatusHistory.objects.filter(changed_by_id__in=op_ids, changed_at__range=(start, end))) if op_ids else []
    rows = []
    for op in ops:
        r = {
            'operator_id': op.id,
            'operator_name': op.full_name or op.username,
            'total': month_assigned.filter(assigned_operator=op).count(),
            **empty_status_counts(),
            'actions_today': 0,
            'touched_today': 0,
            'actions_month': 0,
            'daily_sale': 0,
            'month': date_info['start_date'][:7],
        }
        touched_today_ids = set()
        for h in today_hist:
            if h.changed_by_id == op.id:
                r['actions_today'] += 1
                touched_today_ids.add(h.lead_id)
                if h.new_status == 'sale':
                    r['daily_sale'] += 1
        r['touched_today'] = len(touched_today_ids)
        for h in month_hist:
            if h.changed_by_id == op.id:
                r['actions_month'] += 1
                if h.new_status in r:
                    r[h.new_status] += 1
        r['conversion'] = calc_conversion(r['sale'], r['total'])
        rows.append(r)
    return ok(rows)


@require_auth('boss', 'admin', 'operator', 'filial_rahbari', 'director', 'director_deputy')
def daily_top_operator(request):
    """Berilgan kunda (odatda bugun) eng ko'p sotuv qilgan operatorni qaytaradi.
    Har kuni soat 17:00dagi 'eng zor operator' animatsiyasi uchun ishlatiladi."""
    today, start, end = day_range(request.GET.get('date'))
    user = request.app_user
    ops_qs = AppUser.objects.filter(role='operator', is_active=True)
    if user.role == 'boss':
        ops_qs = ops_qs.filter(boss=user)
    elif user.role == 'operator':
        ops_qs = ops_qs.filter(boss_id=user.boss_id) if user.boss_id else ops_qs.filter(id=user.id)
    elif user.role == 'filial_rahbari':
        pass
    elif user.role in ('director', 'director_deputy'):
        pass
    op_ids = list(ops_qs.values_list('id', flat=True))
    if not op_ids:
        return ok({'date': today, 'top_operator': None})
    sale_rows = (
        LeadStatusHistory.objects
        .filter(changed_by_id__in=op_ids, new_status='sale', changed_at__range=(start, end))
        .values('changed_by_id')
        .annotate(sale_count=Count('id'))
        .order_by('-sale_count')
    )
    sale_rows = list(sale_rows)
    if not sale_rows or not sale_rows[0]['sale_count']:
        return ok({'date': today, 'top_operator': None})
    top = sale_rows[0]
    top_user = AppUser.objects.filter(id=top['changed_by_id']).first()
    return ok({
        'date': today,
        'top_operator': {
            'operator_id': top['changed_by_id'],
            'operator_name': (top_user.full_name or top_user.username) if top_user else 'Operator',
            'sale_count': top['sale_count'],
        },
    })


@require_auth('operator')
def operator_daily_results(request):
    user = request.app_user
    today, start, end = day_range()
    current_month = timezone.localdate().strftime('%Y-%m')
    date_info = parse_date_range({'month': current_month})
    hist_today = LeadStatusHistory.objects.filter(changed_by=user, changed_at__range=(start, end))
    hist_month = LeadStatusHistory.objects.filter(changed_by=user, changed_at__range=(date_info['start_dt'], date_info['end_dt']))
    data = {st: hist_month.filter(new_status=st).count() for st in STATUS_ORDER}
    data['new'] = Lead.objects.filter(assigned_operator=user, created_at__range=(date_info['start_dt'], date_info['end_dt'])).count()
    data.update({
        'current_month': current_month,
        'daily_sale': hist_today.filter(new_status='sale').count(),
        'daily_otkaz': hist_today.filter(new_status='otkaz').count(),
        'daily_advice': hist_today.filter(new_status='advice').count(),
        'daily_open_number': hist_today.filter(new_status='open_number').count(),
        'daily_wrong_number': hist_today.filter(new_status='wrong_number').count(),
        'daily_not_answered': hist_today.filter(new_status='not_answered').count(),
        'touched_today': hist_today.values('lead_id').distinct().count(),
        'actions_today': hist_today.count(),
        'actions_month': hist_month.count(),
    })
    return ok(data)


@require_auth('operator')
def operator_daily_history(request):
    days = int(request.GET.get('days') or 31)
    end = timezone.now()
    start = end - timedelta(days=days)
    hist = LeadStatusHistory.objects.filter(changed_by=request.app_user, changed_at__gte=start).order_by('-changed_at')
    by_day = defaultdict(lambda: {'date': '', 'sale': 0, 'otkaz': 0, 'wrong_number': 0, 'open_number': 0, 'advice': 0, 'other': 0, 'not_answered': 0, 'actions_total': 0})
    for h in hist:
        key = local_date_key(h.changed_at)
        row = by_day[key]
        row['date'] = key
        row['actions_total'] += 1
        if h.new_status in row:
            row[h.new_status] += 1
    return ok({
        'results': [by_day[k] for k in sorted(by_day.keys(), reverse=True)],
        'monthly_archive': build_monthly_archive_for_operators([request.app_user.id], request.GET),
    })


@require_auth('operator')
def operator_monthly_report(request):
    return ok({'results': build_monthly_archive_for_operators([request.app_user.id], request.GET)})


def _users_for_admin_report(params):
    qs = AppUser.objects.select_related('boss').filter(role='operator')
    if clean_string(params.get('boss_id')):
        qs = qs.filter(boss_id=int(params.get('boss_id')))
    if clean_string(params.get('operator_id')):
        qs = qs.filter(id=int(params.get('operator_id')))
    return list(qs.order_by('boss__full_name', 'full_name', 'username'))


def build_admin_report(params):
    date_info = parse_date_range(params)
    operators = _users_for_admin_report(params)
    op_ids = [op.id for op in operators]
    if not op_ids:
        return {'range': {'start_date': date_info['start_date'], 'end_date': date_info['end_date']}, 'selected_operator': {'id': None, 'name': ''}, 'summary': {'assigned_leads': 0, **empty_status_counts(), 'actions_total': 0, 'online_assigned': 0}, 'operators': [], 'operator_rows': [], 'daily': [], 'monthly_archive': []}
    assigned_qs = Lead.objects.filter(assigned_operator_id__in=op_ids, created_at__range=(date_info['start_dt'], date_info['end_dt']))
    hist = list(LeadStatusHistory.objects.select_related('changed_by').filter(changed_by_id__in=op_ids, changed_at__range=(date_info['start_dt'], date_info['end_dt'])))
    online = OnlineLead.objects.filter(assigned_operator_id__in=op_ids, assigned_at__range=(date_info['start_dt'], date_info['end_dt']))
    summary = {'assigned_leads': assigned_qs.count(), **empty_status_counts(), 'actions_total': len(hist), 'online_assigned': online.count()}
    for h in hist:
        if h.new_status in summary:
            summary[h.new_status] += 1
    rows = []
    for op in operators:
        r = {'boss_id': op.boss_id, 'boss_name': op.boss.full_name if op.boss else '', 'operator_id': op.id, 'id': op.id, 'name': op.full_name or op.username, 'operator_name': op.full_name or op.username, 'assigned_leads': assigned_qs.filter(assigned_operator=op).count(), **empty_status_counts(), 'actions_total': 0, 'online_assigned': online.filter(assigned_operator=op).count()}
        for h in hist:
            if h.changed_by_id == op.id:
                r['actions_total'] += 1
                if h.new_status in r:
                    r[h.new_status] += 1
        rows.append(r)
    daily_map = defaultdict(lambda: {'date': '', 'assigned_leads': 0, 'sale': 0, 'otkaz': 0, 'wrong_number': 0, 'open_number': 0, 'advice': 0, 'other': 0, 'not_answered': 0, 'actions_total': 0, 'online_assigned': 0})
    for lead in assigned_qs:
        key = local_date_key(lead.created_at); daily_map[key]['date'] = key; daily_map[key]['assigned_leads'] += 1
    for h in hist:
        key = local_date_key(h.changed_at); daily_map[key]['date'] = key; daily_map[key]['actions_total'] += 1
        if h.new_status in daily_map[key]: daily_map[key][h.new_status] += 1
    for item in online:
        key = local_date_key(item.assigned_at); daily_map[key]['date'] = key; daily_map[key]['online_assigned'] += 1
    selected = operators[0] if len(operators) == 1 else None
    return {'range': {'start_date': date_info['start_date'], 'end_date': date_info['end_date']}, 'selected_operator': {'id': selected.id if selected else None, 'name': (selected.full_name or selected.username) if selected else ''}, 'summary': summary, 'operators': [{'id': r['operator_id'], 'name': r['operator_name'], 'boss_name': r['boss_name']} for r in rows], 'operator_rows': rows, 'daily': [daily_map[k] for k in sorted(daily_map.keys(), reverse=True)], 'monthly_archive': build_monthly_archive_for_operators(op_ids, params, include_online_submitted=True, include_decisions=True)}


@require_auth('admin')
def admin_operator_report(request):
    return ok(build_admin_report(request.GET))


@require_auth('admin')
def admin_monthly_report(request):
    operators = _users_for_admin_report(request.GET)
    op_ids = [op.id for op in operators]
    return ok({'results': build_monthly_archive_for_operators(op_ids, request.GET, include_online_submitted=True, include_decisions=True)})


def build_boss_full_report(user, params):
    query_params = dict(params.items()) if hasattr(params, 'items') else dict(params)
    operator_id = clean_string(query_params.get('operator_id'))
    date_info = parse_date_range(query_params)
    ops = list(AppUser.objects.filter(role='operator', boss=user, is_active=True).order_by('full_name', 'username'))
    if operator_id:
        ops = [op for op in ops if op.id == int(operator_id)]
    op_ids = [op.id for op in ops]
    assigned_qs = Lead.objects.filter(assigned_operator_id__in=op_ids, created_at__range=(date_info['start_dt'], date_info['end_dt'])) if op_ids else Lead.objects.none()
    hist = list(LeadStatusHistory.objects.filter(changed_by_id__in=op_ids, changed_at__range=(date_info['start_dt'], date_info['end_dt']))) if op_ids else []
    online_submitted = OnlineLead.objects.filter(assigned_boss=user, submitted_at__range=(date_info['start_dt'], date_info['end_dt']))
    online_assigned = OnlineLead.objects.filter(assigned_operator_id__in=op_ids, assigned_at__range=(date_info['start_dt'], date_info['end_dt'])) if op_ids else OnlineLead.objects.none()
    decisions = list(LeadVisitDecision.objects.select_related('decided_by', 'payment_done_by', 'payment_not_done_by', 'left_without_payment_by', 'lead', 'lead__assigned_operator', 'lead__boss').filter(lead__assigned_operator_id__in=op_ids, updated_at__range=(date_info['start_dt'], date_info['end_dt']))) if op_ids else []
    summary = {
        'assigned_leads': assigned_qs.count(),
        **empty_status_counts(),
        'actions_total': len(hist),
        'online_submitted': online_submitted.count(),
        'online_assigned': online_assigned.count(),
        'arrived': sum(1 for d in decisions if d.decision == 'arrived'),
        'not_arrived': sum(1 for d in decisions if d.decision == 'not_arrived'),
        'payment_done': sum(1 for d in decisions if visit_payment_status(d) == 'paid'),
        'payment_not_done': sum(1 for d in decisions if visit_payment_status(d) == 'unpaid'),
        'left_without_payment': sum(1 for d in decisions if visit_payment_status(d) == 'left_without_payment'),
        'payment_pending': sum(1 for d in decisions if visit_payment_status(d) == 'pending'),
    }
    for h in hist:
        if h.new_status in summary: summary[h.new_status] += 1
    op_rows = []
    for op in ops:
        r = {'operator_id': op.id, 'operator_name': op.full_name or op.username, 'assigned_leads': assigned_qs.filter(assigned_operator=op).count(), 'online_assigned': online_assigned.filter(assigned_operator=op).count(), **empty_status_counts(), 'actions_total': 0}
        for h in hist:
            if h.changed_by_id == op.id:
                r['actions_total'] += 1
                if h.new_status in r: r[h.new_status] += 1
        op_rows.append(r)
    filial_map = {}
    for d in decisions:
        key = d.decided_by_id or 0
        filial_map.setdefault(key, {
            'filial_rahbari_id': d.decided_by_id,
            'filial_rahbari_name': d.decided_by.full_name if d.decided_by else '-',
            'arrived': 0,
            'not_arrived': 0,
            'payment_done': 0,
            'payment_not_done': 0,
            'left_without_payment': 0,
            'payment_pending': 0,
            'total': 0,
        })
        filial_map[key]['total'] += 1
        filial_map[key]['arrived' if d.decision == 'arrived' else 'not_arrived'] += 1
        payment_key = {
            'paid': 'payment_done',
            'unpaid': 'payment_not_done',
            'left_without_payment': 'left_without_payment',
            'pending': 'payment_pending',
        }[visit_payment_status(d)]
        filial_map[key][payment_key] += 1
    daily_map = defaultdict(lambda: {
        'date': '', 'assigned_leads': 0, 'sale': 0, 'otkaz': 0,
        'wrong_number': 0, 'open_number': 0, 'advice': 0, 'other': 0,
        'not_answered': 0, 'online_submitted': 0, 'online_assigned': 0,
        'arrived': 0, 'not_arrived': 0, 'payment_done': 0,
        'payment_not_done': 0, 'left_without_payment': 0, 'payment_pending': 0,
    })
    for lead in assigned_qs:
        key = local_date_key(lead.created_at); daily_map[key]['date'] = key; daily_map[key]['assigned_leads'] += 1
    for h in hist:
        key = local_date_key(h.changed_at); daily_map[key]['date'] = key
        if h.new_status in daily_map[key]: daily_map[key][h.new_status] += 1
    for o in online_submitted:
        key = local_date_key(o.submitted_at); daily_map[key]['date'] = key; daily_map[key]['online_submitted'] += 1
    for o in online_assigned:
        key = local_date_key(o.assigned_at); daily_map[key]['date'] = key; daily_map[key]['online_assigned'] += 1
    for d in decisions:
        key = local_date_key(d.updated_at)
        daily_map[key]['date'] = key
        daily_map[key]['arrived' if d.decision == 'arrived' else 'not_arrived'] += 1
        payment_key = {
            'paid': 'payment_done',
            'unpaid': 'payment_not_done',
            'left_without_payment': 'left_without_payment',
            'pending': 'payment_pending',
        }[visit_payment_status(d)]
        daily_map[key][payment_key] += 1
    selected = ops[0] if len(ops) == 1 else None
    decision_rows = [visit_decision_to_dict(d) for d in sorted(decisions, key=lambda x: x.updated_at, reverse=True)]
    return {
        'range': {'start_date': date_info['start_date'], 'end_date': date_info['end_date']},
        'selected_operator': {
            'id': selected.id if selected else None,
            'name': (selected.full_name or selected.username) if selected else '',
        },
        'summary': summary,
        'operators': op_rows,
        'filial_rahbarlari': list(filial_map.values()),
        'daily': [daily_map[k] for k in sorted(daily_map.keys(), reverse=True)],
        'monthly_archive': build_monthly_archive_for_operators(
            op_ids, params, boss_user=user,
            include_online_submitted=True, include_decisions=True,
        ),
        'visit_decisions': decision_rows,
        'not_arrived_leads': [row for row in decision_rows if row.get('decision') == 'not_arrived'],
        'arrived_leads': [row for row in decision_rows if row.get('decision') == 'arrived'],
        'payment_done_leads': [row for row in decision_rows if row.get('payment_status') == 'paid'],
        'payment_not_done_leads': [row for row in decision_rows if row.get('payment_status') == 'unpaid'],
        'left_without_payment_leads': [row for row in decision_rows if row.get('payment_status') == 'left_without_payment'],
        'payment_pending_leads': [row for row in decision_rows if row.get('payment_status') == 'pending'],
    }


@require_auth('boss')
def boss_full_report(request):
    return ok(build_boss_full_report(request.app_user, request.GET))


@require_auth('boss')
def boss_monthly_report(request):
    ops = list(AppUser.objects.filter(role='operator', boss=request.app_user, is_active=True).order_by('full_name', 'username'))
    operator_id = clean_string(request.GET.get('operator_id'))
    if operator_id:
        ops = [op for op in ops if str(op.id) == operator_id]
    op_ids = [op.id for op in ops]
    return ok({'results': build_monthly_archive_for_operators(op_ids, request.GET, boss_user=request.app_user, include_online_submitted=True, include_decisions=True)})


@require_auth('boss')
def boss_accounting_report(request):
    # Operatorlar boshligi uchun Hisob kitob: oy bo‘yicha operator sotuvlari va oyma-oy sotuv arxivi.
    selected_month_date = parse_month_key(request.GET.get('month')) or timezone.localdate().replace(day=1)
    start_dt = timezone.make_aware(datetime.combine(selected_month_date, time.min))
    end_dt = timezone.make_aware(datetime.combine(add_months(selected_month_date, 1), time.min))

    # Hisob-kitob tarixiy bo‘lgani uchun inactive operatorlar ham chiqadi.
    operators = list(AppUser.objects.filter(role='operator', boss=request.app_user).order_by('full_name', 'username'))
    op_ids = [op.id for op in operators]

    sales_count_map = defaultdict(int)
    unique_sales_map = defaultdict(int)
    if op_ids:
        grouped_sales = LeadStatusHistory.objects.filter(
            changed_by_id__in=op_ids,
            new_status='sale',
            changed_at__gte=start_dt,
            changed_at__lt=end_dt,
        ).values('changed_by_id').annotate(sale_count=Count('id'), unique_sales=Count('lead_id', distinct=True))
        for row in grouped_sales:
            sales_count_map[row['changed_by_id']] = row['sale_count'] or 0
            unique_sales_map[row['changed_by_id']] = row['unique_sales'] or 0

    operator_rows = []
    for op in operators:
        operator_rows.append({
            'operator_id': op.id,
            'operator_name': op.full_name or op.username,
            'username': op.username,
            'is_active': bool(op.is_active),
            'sale': sales_count_map.get(op.id, 0),
            'unique_sales': unique_sales_map.get(op.id, 0),
        })
    operator_rows.sort(key=lambda item: (-item['sale'], item['operator_name']))

    period_params = dict(request.GET.items()) if hasattr(request.GET, 'items') else {}
    period_params.setdefault('end_month', month_key_from_date(selected_month_date))
    period_params.setdefault('months', '12')
    monthly_archive = build_monthly_archive_for_operators(
        op_ids,
        period_params,
        boss_user=request.app_user,
        include_online_submitted=True,
        include_decisions=True,
    )

    month_options = []
    cursor = selected_month_date
    for _ in range(12):
        key = month_key_from_date(cursor)
        month_options.append({'value': key, 'label': month_label(key)})
        cursor = add_months(cursor, -1)

    total_sale = sum(row['sale'] for row in operator_rows)
    top_operator = operator_rows[0] if operator_rows else None
    return ok({
        'selected_month': month_key_from_date(selected_month_date),
        'selected_month_label': month_label(month_key_from_date(selected_month_date)),
        'month_options': month_options,
        'summary': {
            'operators_count': len(operator_rows),
            'total_sale': total_sale,
            'active_operators': sum(1 for op in operators if op.is_active),
            'top_operator': top_operator['operator_name'] if top_operator and top_operator.get('sale') else '',
        },
        'operators': operator_rows,
        'monthly_archive': monthly_archive,
    })


def add_header_style(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E78')
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'
    autosize(ws)


def _style_daily_report_sheet(ws, title_row=1):
    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(bold=True, color='FFFFFF')
    total_fill = PatternFill('solid', fgColor='D9EAF7')
    for cell in ws[title_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.freeze_panes = f'A{title_row + 1}'
    for row in ws.iter_rows(min_row=title_row + 1):
        if row and row[0].value and str(row[0].value).startswith('Bugungi umumiy'):
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = total_fill
    autosize(ws)


def _append_daily_sale_lead_header(ws):
    ws.append([
        'Vaqt', 'Operator', 'F.I.O', 'T/SH', 'Maktab', 'Sinf', 'Fan', 'Ball',
        'tel1', 'tel2', 'tel3', 'Filial', 'Oldingi holat', 'Yangi holat', 'Izoh', 'Lead ID'
    ])


def _append_daily_sale_lead_row(ws, history):
    lead = history.lead
    operator = history.changed_by or (lead.assigned_operator if lead else None)
    ws.append([
        format_dt(history.changed_at),
        operator.full_name if operator else '',
        excel_safe(lead.full_name if lead else ''),
        excel_safe(lead.tsh if lead else ''),
        excel_safe(lead.school if lead else ''),
        excel_safe(lead.grade if lead else ''),
        excel_safe(lead.subject if lead else ''),
        excel_safe(lead.ball if lead else ''),
        excel_safe(lead.phone1 if lead else ''),
        excel_safe(lead.phone2 if lead else ''),
        excel_safe(lead.phone3 if lead else ''),
        excel_safe(lead.branch_name if lead else ''),
        STATUS_LABELS.get(history.old_status, history.old_status),
        STATUS_LABELS.get(history.new_status, history.new_status),
        excel_safe(history.note),
        history.lead_id,
    ])


@require_auth('boss')
def boss_daily_report_excel(request):
    selected_date, start, end = day_range(request.GET.get('date'))
    operators = list(AppUser.objects.filter(role='operator', boss=request.app_user).order_by('full_name', 'username'))
    op_ids = [op.id for op in operators]

    histories = list(
        LeadStatusHistory.objects
        .select_related('lead', 'lead__assigned_operator', 'changed_by')
        .filter(changed_by_id__in=op_ids, changed_at__range=(start, end))
        .order_by('changed_by__full_name', 'changed_by__username', 'changed_at')
    ) if op_ids else []

    status_keys = ['sale', 'open_number', 'advice', 'otkaz', 'wrong_number', 'other', 'not_answered']
    operator_stats = {op.id: {key: 0 for key in status_keys} for op in operators}
    for history in histories:
        if history.changed_by_id in operator_stats and history.new_status in operator_stats[history.changed_by_id]:
            operator_stats[history.changed_by_id][history.new_status] += 1

    wb = Workbook()
    ws = wb.active
    ws.title = 'Kunlik hisobot'
    ws.append([
        'Operator ism familyasi', 'Sotuv', "O'chiq Nomer", 'Maslahat', 'Atkaz',
        'Xato nomer', "O'qiydi", "Ko'tarmadi", "Umumiy bog'langanlar",
        "Qo'shimcha vazifa", "Qo'shimcha vazifa vaqti"
    ])

    total_sales = 0
    for op in operators:
        counts = operator_stats.get(op.id, {})
        total_sales += counts.get('sale', 0)
        connected_total = sum(counts.get(key, 0) for key in status_keys)
        ws.append([
            op.full_name or op.username,
            counts.get('sale', 0),
            counts.get('open_number', 0),
            counts.get('advice', 0),
            counts.get('otkaz', 0),
            counts.get('wrong_number', 0),
            counts.get('other', 0),
            counts.get('not_answered', 0),
            connected_total,
            '',
            '',
        ])

    ws.append([])
    ws.append(['Tanlangan kundagi umumiy sotuvlar soni', total_sales])
    ws.append(['Hisobot sanasi', selected_date])
    ws.append(['Boshliq', request.app_user.full_name or request.app_user.username])
    _style_daily_report_sheet(ws)

    sales_histories = [history for history in histories if history.new_status == 'sale']
    by_operator = defaultdict(list)
    for history in sales_histories:
        by_operator[history.changed_by_id].append(history)

    all_sales = wb.create_sheet('Barcha sotuvlar')
    _append_daily_sale_lead_header(all_sales)
    for history in sales_histories:
        _append_daily_sale_lead_row(all_sales, history)
    _style_daily_report_sheet(all_sales)

    used_titles = set([ws.title, all_sales.title])
    for op in operators:
        base_title = _safe_sheet_title(op.full_name or op.username, 'Operator')
        title = base_title
        counter = 2
        while title in used_titles:
            suffix = f' {counter}'
            title = (base_title[:31 - len(suffix)] + suffix).strip()
            counter += 1
        used_titles.add(title)
        sheet = wb.create_sheet(title)
        _append_daily_sale_lead_header(sheet)
        for history in by_operator.get(op.id, []):
            _append_daily_sale_lead_row(sheet, history)
        if not by_operator.get(op.id):
            sheet.append(['Bugun bu operator sotuv qilmagan'])
        _style_daily_report_sheet(sheet)

    filename = f'kunlik_hisobot_{selected_date}.xlsx'
    return excel_response(wb, filename)


@require_auth('boss')
def boss_full_report_excel(request):
    report = build_boss_full_report(request.app_user, request.GET)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Natijalar'
    ws.append(['Ko‘rsatkich', 'Soni'])
    summary_labels = {
        'assigned_leads': 'Biriktirilgan leadlar',
        'sale': 'Sotuv',
        'otkaz': 'Atkaz',
        'wrong_number': 'Xato nomer',
        'open_number': "O‘chiq nomer",
        'advice': 'Maslahat',
        'other': "O‘qiydi",
        'not_answered': "Ko‘tarmadi",
        'actions_total': 'Jami harakatlar',
        'online_submitted': 'Online yuborilgan',
        'online_assigned': 'Online biriktirilgan',
        'arrived': 'Keldi',
        'not_arrived': 'Kelmadi',
        'payment_done': 'To‘lov qildi',
        'payment_not_done': 'To‘lov qilmadi',
        'left_without_payment': 'Keldi, to‘lov qilmasdan ketdi',
        'payment_pending': 'To‘lov belgilanmagan',
    }
    for key, val in report['summary'].items():
        ws.append([summary_labels.get(key, key), val])
    ws.append([])
    ws.append([
        'Operator', 'Biriktirildi', 'Online', 'Sotuv', 'Atkaz', 'Xato nomer',
        "O‘chiq nomer", 'Maslahat', "O‘qiydi", "Ko‘tarmadi", 'Jami action',
    ])
    for row in report['operators']:
        ws.append([
            row['operator_name'], row['assigned_leads'], row['online_assigned'],
            row['sale'], row['otkaz'], row['wrong_number'], row['open_number'],
            row['advice'], row['other'], row.get('not_answered', 0), row['actions_total'],
        ])
    add_header_style(ws)

    ws2 = wb.create_sheet('Menenjerlar')
    ws2.append([
        'Menenjer', 'Jami', 'Keldi', 'Kelmadi',
        'To‘lov qildi', 'To‘lov qilmadi', 'Keldi, to‘lov qilmasdan ketdi', 'To‘lov belgilanmagan',
    ])
    for row in report.get('filial_rahbarlari', []):
        ws2.append([
            row.get('filial_rahbari_name', ''), row.get('total', 0),
            row.get('arrived', 0), row.get('not_arrived', 0),
            row.get('payment_done', 0), row.get('payment_not_done', 0),
            row.get('left_without_payment', 0), row.get('payment_pending', 0),
        ])
    add_header_style(ws2)

    def append_decision_sheet(sheet_name, rows):
        sheet = wb.create_sheet(sheet_name)
        sheet.append([
            'Belgi vaqti', 'Menenjer', 'Filial', 'Qaror', 'To‘lov holati',
            'To‘lov holatini belgilagan', 'To‘lov holati vaqti', 'F.I.O',
            'tel1', 'tel2', 'tel3', 'T/SH', 'Maktab', 'Sinf', 'Fan',
            'Ball', 'Operator',
        ])
        for item in rows:
            sheet.append([
                format_dt(item.get('updated_at')),
                item.get('filial_rahbari_name', ''),
                item.get('filial_rahbari_branch', ''),
                'Keldi' if item.get('decision') == 'arrived' else 'Kelmadi',
                {
                    'paid': 'To‘lov qildi',
                    'unpaid': 'To‘lov qilmadi',
                    'left_without_payment': 'Keldi, to‘lov qilmasdan ketdi',
                    'pending': 'Belgilanmagan',
                }.get(item.get('payment_status'), 'Belgilanmagan'),
                item.get('payment_status_by_name') or '',
                format_dt(item.get('payment_status_at')),
                excel_safe(item.get('lead_name') or item.get('full_name') or ''),
                item.get('lead_phone') or item.get('phone1') or '',
                item.get('lead_phone2') or item.get('phone2') or '',
                item.get('lead_phone3') or item.get('phone3') or '',
                item.get('tsh') or '',
                item.get('display_school') or item.get('school') or '',
                item.get('grade') or '',
                item.get('subject') or '',
                item.get('ball') or '',
                item.get('operator_name') or '',
            ])
        add_header_style(sheet)

    append_decision_sheet('Barcha Nazorat', report.get('visit_decisions', []))
    append_decision_sheet('Keldi Leadlar', report.get('arrived_leads', []))
    append_decision_sheet('Kelmadi Leadlar', report.get('not_arrived_leads', []))
    append_decision_sheet('To‘lov Qilganlar', report.get('payment_done_leads', []))
    append_decision_sheet('To‘lov Qilmaganlar', report.get('payment_not_done_leads', []))
    append_decision_sheet('Tolovsiz Ketganlar', report.get('left_without_payment_leads', []))
    append_decision_sheet('To‘lov Belgilanmagan', report.get('payment_pending_leads', []))
    return excel_response(wb, 'umumiy_hisobot.xlsx')


def _safe_sheet_title(value, fallback='Operator'):
    title = excel_safe(value or fallback).strip() or fallback
    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
        title = title.replace(ch, ' ')
    return title[:31]


def _split_datetime_parts(dt):
    if not dt:
        return ['', '', '', '']
    local_dt = timezone.localtime(dt) if timezone.is_aware(dt) else dt
    return [local_dt.year, local_dt.month, local_dt.day, local_dt.strftime('%H:%M')]


def _append_monthly_sales_header(ws):
    ws.append([
        'ID', 'Yil', 'Oy', 'Kun', 'Soat', 'Operator', 'Boshliq', 'Filial',
        'F.I.O', 'T/SH', 'Maktab', 'Sinf', 'Fan', 'Ball', 'tel1', 'tel2', 'tel3',
        'Oldingi status', 'Yangi status', 'Izoh'
    ])


def _append_monthly_sales_row(ws, h):
    lead = h.lead
    operator = h.changed_by or (lead.assigned_operator if lead else None)
    boss = lead.boss if lead else None
    year, month, day, hour = _split_datetime_parts(h.changed_at)
    ws.append([
        h.id,
        year,
        month,
        day,
        hour,
        operator.full_name if operator else '',
        boss.full_name if boss else '',
        lead.branch_name if lead else '',
        excel_safe(lead.full_name if lead else ''),
        excel_safe(lead.tsh if lead else ''),
        excel_safe(lead.school if lead else ''),
        excel_safe(lead.grade if lead else ''),
        excel_safe(lead.subject if lead else ''),
        excel_safe(lead.ball if lead else ''),
        excel_safe(lead.phone1 if lead else ''),
        excel_safe(lead.phone2 if lead else ''),
        excel_safe(lead.phone3 if lead else ''),
        STATUS_LABELS.get(h.old_status, h.old_status),
        STATUS_LABELS.get(h.new_status, h.new_status),
        excel_safe(h.note),
    ])


@require_auth('admin')
def admin_monthly_sales_excel(request):
    month = clean_string(request.GET.get('month'))
    parsed_month = parse_month_key(month)
    if not parsed_month:
        return JsonResponse({'detail': 'Oy noto‘g‘ri tanlangan. Masalan: 2026-06'}, status=400)

    next_month = add_months(parsed_month, 1)
    start_dt = timezone.make_aware(datetime.combine(parsed_month, time.min))
    end_dt = timezone.make_aware(datetime.combine(next_month - timedelta(days=1), time.max))

    sales = list(
        LeadStatusHistory.objects
        .select_related('lead', 'lead__assigned_operator', 'lead__boss', 'changed_by')
        .filter(new_status='sale', changed_at__range=(start_dt, end_dt))
        .order_by('changed_by__full_name', 'changed_by__username', 'changed_at')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Barcha sotuvlar'
    _append_monthly_sales_header(ws)
    for h in sales:
        _append_monthly_sales_row(ws, h)
    add_header_style(ws)

    operator_sales = defaultdict(list)
    for h in sales:
        lead = h.lead
        operator = h.changed_by or (lead.assigned_operator if lead else None)
        key = operator.id if operator else 0
        operator_sales[key].append(h)

    used_titles = {'Barcha sotuvlar'}
    for index, (operator_id, rows) in enumerate(operator_sales.items(), start=1):
        first = rows[0]
        lead = first.lead
        operator = first.changed_by or (lead.assigned_operator if lead else None)
        base_title = _safe_sheet_title(operator.full_name if operator else 'Operatorsiz', f'Operator {index}')
        title = base_title
        counter = 2
        while title in used_titles:
            suffix = f' {counter}'
            title = f'{base_title[:31-len(suffix)]}{suffix}'
            counter += 1
        used_titles.add(title)

        sheet = wb.create_sheet(title)
        _append_monthly_sales_header(sheet)
        for h in rows:
            _append_monthly_sales_row(sheet, h)
        add_header_style(sheet)

    filename = f"oylik_sotuvlar_{parsed_month.strftime('%Y_%m')}.xlsx"
    return excel_response(wb, filename)


def _apply_visit_decision_report_filters(qs, params):
    if any(clean_string(params.get(key)) for key in ('date', 'month', 'start_date', 'end_date')):
        date_info = parse_date_range(params)
        qs = qs.filter(updated_at__range=(date_info['start_dt'], date_info['end_dt']))

    manager_id = clean_string(params.get('manager_id') or params.get('filial_rahbari_id'))
    if manager_id and manager_id.isdigit():
        qs = qs.filter(decided_by_id=int(manager_id))

    decision = clean_string(params.get('decision'))
    if decision in ('arrived', 'not_arrived'):
        qs = qs.filter(decision=decision)

    payment = clean_string(params.get('payment'))
    if payment == 'paid':
        qs = qs.filter(payment_done=True)
    elif payment == 'unpaid':
        qs = qs.filter(payment_not_done=True)
    elif payment == 'left_without_payment':
        qs = qs.filter(left_without_payment=True)
    elif payment == 'pending':
        qs = qs.filter(payment_done=False, payment_not_done=False, left_without_payment=False)
    return qs


def visit_decisions_excel_response(qs, filename):
    decisions = list(qs.order_by('-updated_at', '-id'))
    paid = [item for item in decisions if visit_payment_status(item) == 'paid']
    unpaid = [item for item in decisions if visit_payment_status(item) == 'unpaid']
    left_without_payment = [item for item in decisions if visit_payment_status(item) == 'left_without_payment']
    pending = [item for item in decisions if visit_payment_status(item) == 'pending']

    wb = Workbook()
    summary = wb.active
    summary.title = 'Statistika'
    summary.append(['Ko‘rsatkich', 'Soni'])
    summary.append(['Jami nazorat', len(decisions)])
    summary.append(['Keldi', sum(1 for item in decisions if item.decision == 'arrived')])
    summary.append(['Kelmadi', sum(1 for item in decisions if item.decision == 'not_arrived')])
    summary.append(['To‘lov qildi', len(paid)])
    summary.append(['To‘lov qilmadi', len(unpaid)])
    summary.append(['Keldi, to‘lov qilmasdan ketdi', len(left_without_payment)])
    summary.append(['To‘lov belgilanmagan', len(pending)])
    add_header_style(summary)

    manager_rows = defaultdict(lambda: {
        'name': '', 'branch': '', 'total': 0, 'arrived': 0, 'not_arrived': 0,
        'paid': 0, 'unpaid': 0, 'left_without_payment': 0, 'pending': 0,
    })
    for item in decisions:
        row = manager_rows[item.decided_by_id or 0]
        row['name'] = (item.decided_by.full_name or item.decided_by.username) if item.decided_by else '-'
        row['branch'] = item.decided_by.branch_name if item.decided_by else ''
        row['total'] += 1
        row['arrived' if item.decision == 'arrived' else 'not_arrived'] += 1
        row[visit_payment_status(item)] += 1

    managers = wb.create_sheet('Menenjerlar')
    managers.append([
        'Menenjer', 'Filial', 'Jami', 'Keldi', 'Kelmadi',
        'To‘lov qildi', 'To‘lov qilmadi', 'Keldi, to‘lov qilmasdan ketdi', 'Belgilanmagan',
    ])
    for row in sorted(manager_rows.values(), key=lambda value: (-value['total'], value['name'])):
        managers.append([
            row['name'], row['branch'], row['total'], row['arrived'], row['not_arrived'],
            row['paid'], row['unpaid'], row['left_without_payment'], row['pending'],
        ])
    add_header_style(managers)

    def add_rows(sheet_name, items):
        sheet = wb.create_sheet(sheet_name)
        sheet.append([
            'Belgi vaqti', 'Menenjer', 'Filial', 'Qaror', 'To‘lov holati',
            'To‘lov holatini belgilagan', 'To‘lov holati vaqti', 'F.I.O',
            'tel1', 'tel2', 'tel3', 'T/SH', 'Maktab', 'Sinf', 'Fan',
            'Ball', 'Operator', 'Boshliq',
        ])
        for item in items:
            lead = item.lead
            sheet.append([
                format_dt(item.updated_at),
                (item.decided_by.full_name or item.decided_by.username) if item.decided_by else '',
                item.decided_by.branch_name if item.decided_by else '',
                'Keldi' if item.decision == 'arrived' else 'Kelmadi',
                visit_payment_label(item),
                visit_payment_status_by_name(item),
                format_dt(visit_payment_status_at(item)),
                excel_safe(lead.full_name if lead else ''),
                excel_safe(lead.phone1 if lead else ''),
                excel_safe(lead.phone2 if lead else ''),
                excel_safe(lead.phone3 if lead else ''),
                excel_safe(lead.tsh if lead else ''),
                excel_safe(lead.school if lead else ''),
                excel_safe(lead.grade if lead else ''),
                excel_safe(lead.subject if lead else ''),
                excel_safe(lead.ball if lead else ''),
                (lead.assigned_operator.full_name or lead.assigned_operator.username) if lead and lead.assigned_operator else '',
                (lead.boss.full_name or lead.boss.username) if lead and lead.boss else '',
            ])
        add_header_style(sheet)

    add_rows('Barcha Natijalar', decisions)
    add_rows('Keldi', [item for item in decisions if item.decision == 'arrived'])
    add_rows('Kelmadi', [item for item in decisions if item.decision == 'not_arrived'])
    add_rows('To‘lov Qildi', paid)
    add_rows('To‘lov Qilmadi', unpaid)
    add_rows('Tolovsiz Ketdi', left_without_payment)
    add_rows('Belgilanmagan', pending)
    return excel_response(wb, filename)


@require_auth('boss')
def boss_visit_decisions_excel(request):
    operator_ids = list(
        AppUser.objects.filter(role='operator', boss=request.app_user).values_list('id', flat=True)
    )
    qs = LeadVisitDecision.objects.select_related(
        'lead', 'lead__assigned_operator', 'lead__boss', 'decided_by',
        'payment_done_by', 'payment_not_done_by', 'left_without_payment_by',
    ).filter(lead__assigned_operator_id__in=operator_ids)
    qs = _apply_visit_decision_report_filters(qs, request.GET)
    return visit_decisions_excel_response(qs, 'boss_keldi_kelmadi_tolov_hisoboti.xlsx')


@require_auth('admin')
def admin_visit_decisions_excel(request):
    qs = LeadVisitDecision.objects.select_related(
        'lead', 'lead__assigned_operator', 'lead__boss', 'decided_by',
        'payment_done_by', 'payment_not_done_by', 'left_without_payment_by',
    )
    qs = _apply_visit_decision_report_filters(qs, request.GET)
    return visit_decisions_excel_response(qs, 'admin_keldi_kelmadi_tolov_hisoboti.xlsx')


@require_auth('admin')
def admin_visit_decisions(request):
    qs = LeadVisitDecision.objects.select_related('lead', 'lead__assigned_operator', 'lead__boss', 'decided_by', 'payment_done_by', 'payment_not_done_by', 'left_without_payment_by').order_by('-updated_at')
    return ok([visit_decision_to_dict(x) for x in qs[:1000]])



def _safe_excel_sheet_title(value, used_titles=None):
    used_titles = used_titles if used_titles is not None else set()
    title = clean_string(value) or 'Operatorsiz'
    for ch in ['\\', '/', '*', '?', ':', '[', ']']:
        title = title.replace(ch, '-')
    title = title[:31] or 'Sheet'
    base = title
    index = 2
    while title in used_titles:
        suffix = f' {index}'
        title = f'{base[:31-len(suffix)]}{suffix}'
        index += 1
    used_titles.add(title)
    return title


def _excel_date_parts(value):
    if not value:
        return ['', '', '', '']
    local_value = timezone.localtime(value) if timezone.is_aware(value) else value
    return [local_value.year, local_value.month, local_value.day, local_value.strftime('%H:%M')]


def _append_monthly_sales_header(ws):
    ws.append([
        'ID', 'Yil', 'Oy', 'Kun', 'Soat', 'Operator', 'F.I.O', 'T/SH', 'Maktab', 'Sinf',
        'Fan', 'Ball', 'tel1', 'tel2', 'tel3', 'Filial', 'Oldingi holat', 'Yangi holat', 'Izoh'
    ])


def _append_monthly_sales_row(ws, history):
    lead = history.lead
    operator = history.changed_by or (lead.assigned_operator if lead else None)
    year, month, day, hour = _excel_date_parts(history.changed_at)
    ws.append([
        lead.id if lead else history.lead_id,
        year,
        month,
        day,
        hour,
        operator.full_name if operator and operator.full_name else (operator.username if operator else ''),
        excel_safe(lead.full_name if lead else ''),
        excel_safe(lead.tsh if lead else ''),
        excel_safe(lead.school if lead else ''),
        excel_safe(lead.grade if lead else ''),
        excel_safe(lead.subject if lead else ''),
        excel_safe(lead.ball if lead else ''),
        excel_safe(lead.phone1 if lead else ''),
        excel_safe(lead.phone2 if lead else ''),
        excel_safe(lead.phone3 if lead else ''),
        excel_safe(lead.branch_name if lead else ''),
        STATUS_LABELS.get(history.old_status, history.old_status),
        STATUS_LABELS.get(history.new_status, history.new_status),
        excel_safe(history.note),
    ])


@require_auth('admin')
def admin_monthly_sales_excel(request):
    selected_month_date = parse_month_key(request.GET.get('month')) or timezone.localdate().replace(day=1)
    start_dt = timezone.make_aware(datetime.combine(selected_month_date, time.min))
    end_dt = timezone.make_aware(datetime.combine(add_months(selected_month_date, 1), time.min))

    operators = _users_for_admin_report(request.GET)
    op_ids = [op.id for op in operators]

    histories = []
    if op_ids:
        histories = list(
            LeadStatusHistory.objects
            .select_related('lead', 'lead__assigned_operator', 'changed_by')
            .filter(changed_by_id__in=op_ids, new_status='sale', changed_at__gte=start_dt, changed_at__lt=end_dt)
            .order_by('-changed_at', '-id')
        )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Barcha sotuvlar'
    _append_monthly_sales_header(ws)
    for item in histories:
        _append_monthly_sales_row(ws, item)
    add_header_style(ws)

    used_titles = {'Barcha sotuvlar'}
    for operator in operators:
        operator_histories = [item for item in histories if item.changed_by_id == operator.id]
        sheet_title = _safe_excel_sheet_title(operator.full_name or operator.username or f'Operator {operator.id}', used_titles)
        sheet = wb.create_sheet(sheet_title)
        _append_monthly_sales_header(sheet)
        for item in operator_histories:
            _append_monthly_sales_row(sheet, item)
        add_header_style(sheet)

    filename = f"oylik_sotuvlar_{month_key_from_date(selected_month_date)}.xlsx"
    return excel_response(wb, filename)


@require_auth('admin')
def admin_all_reports_excel(request):
    wb = Workbook()
    ws = wb.active; ws.title = 'Barcha Leadlar'
    ws.append(['F.I.O', 'T/SH', 'Maktab', 'Sinf', 'Fan', 'Ball', 'tel1', 'tel2', 'tel3', 'Status', 'Operator', 'Yangilangan'])
    for l in Lead.objects.select_related('assigned_operator', 'boss').order_by('-id'):
        ws.append([excel_safe(l.full_name), l.tsh, l.school, l.grade, l.subject, l.ball, l.phone1, l.phone2, l.phone3, STATUS_LABELS.get(l.current_status, l.current_status), l.assigned_operator.full_name if l.assigned_operator else '', format_dt(l.updated_at)])
    add_header_style(ws)
    ws2 = wb.create_sheet('Status Tarixi')
    ws2.append(['Vaqt', 'Lead ID', 'F.I.O', 'Operator', 'Oldingi status', 'Yangi status', 'Izoh'])
    for h in LeadStatusHistory.objects.select_related('lead', 'changed_by').order_by('-changed_at'):
        ws2.append([format_dt(h.changed_at), h.lead_id, h.lead.full_name if h.lead else '', h.changed_by.full_name if h.changed_by else '', STATUS_LABELS.get(h.old_status, h.old_status), STATUS_LABELS.get(h.new_status, h.new_status), excel_safe(h.note)])
    add_header_style(ws2)
    ws3 = wb.create_sheet('Online Leadlar')
    ws3.append(['ID', 'F.I.O', 'Yosh', 'tel1', 'tel2', 'tel3', 'Fan', 'Hudud', 'Boss', 'Operator', 'Yuborilgan', 'Biriktirilgan'])
    for o in OnlineLead.objects.select_related('assigned_boss', 'assigned_operator').order_by('-submitted_at'):
        ws3.append([o.id, o.full_name, o.age, o.phone1, o.phone2, o.phone3, o.interest_subject, o.region, o.assigned_boss.full_name if o.assigned_boss else '', o.assigned_operator.full_name if o.assigned_operator else '', format_dt(o.submitted_at), format_dt(o.assigned_at)])
    add_header_style(ws3)
    ws4 = wb.create_sheet('Keldi Kelmadi Tolov')
    ws4.append([
        'Vaqt', 'Lead', 'Telefon', 'Menenjer', 'Filial', 'Qaror',
        'To‘lov holati', 'Holatni belgilagan', 'To‘lov holati vaqti',
    ])
    for d in LeadVisitDecision.objects.select_related(
        'lead', 'decided_by', 'payment_done_by', 'payment_not_done_by', 'left_without_payment_by'
    ).order_by('-updated_at'):
        ws4.append([
            format_dt(d.updated_at),
            d.lead.full_name if d.lead else '',
            d.lead.phone1 if d.lead else '',
            (d.decided_by.full_name or d.decided_by.username) if d.decided_by else '',
            d.decided_by.branch_name if d.decided_by else '',
            'Keldi' if d.decision == 'arrived' else 'Kelmadi',
            visit_payment_label(d),
            visit_payment_status_by_name(d),
            format_dt(visit_payment_status_at(d)),
        ])
    add_header_style(ws4)
    ws5 = wb.create_sheet('Excel Import Qatorlari')
    ws5.append(['Import ID', 'Qator', 'Status', 'Lead ID', 'Xato', 'Raw data'])
    for r in ExcelImportRow.objects.order_by('-id')[:5000]:
        ws5.append([r.import_record_id, r.source_row_number, r.status, r.lead_id or '', r.error_message, json.dumps(r.raw_data, ensure_ascii=False)])
    add_header_style(ws5)
    ws6 = wb.create_sheet('Foydalanuvchilar')
    ws6.append(['ID', 'Login', 'Ism', 'Rol', 'Telefon', 'Boss', 'Filial', 'Aktiv'])
    for u in AppUser.objects.select_related('boss').order_by('role', 'full_name'):
        ws6.append([u.id, u.username, u.full_name, u.role, u.phone, u.boss.full_name if u.boss else '', u.branch_name, 'Ha' if u.is_active else 'Yo‘q'])
    add_header_style(ws6)
    return excel_response(wb, 'barcha_hisobotlar.xlsx')
